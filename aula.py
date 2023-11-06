# -*- coding: utf-8 -*-


'''
EXERCICIOS PRIMEIRO SEMESTRE

'''


"""
Created on Wed Mar 22 10:38:50 2023

@author: Prof Eder

Apostila do curso de Python

FIAP - Engenharia de Software - resolucao da apostila
"""

a = 1
b = 2
c = a + b

#listagem 4.3 - Carro novo ou velho, dependendo da idade.

idade = int(input('Digite a idade do carro:'))

if idade <= 3:
    print('seu carro Ã© novo')

if idade > 3:
    print('seu carro Ã© velho')


#Exemplo de calculo de mÃ©dia 6.5

notas = [6, 7, 5, 8, 9]
soma = 0
x = 0

while x < 5: 
    soma += notas[x]
    x += 1

print("MÃ©dia: %5.2f" %(soma/x))


#Exemplo de calculo de mÃ©dia 6.6

notas = [0,0,0,0,0]
soma = 0
x = 0
while x < 5:
    notas[x] = float(input("Nota %d:" % x)) 
    soma += notas[x]
    x += 1
x = 0 
while x < 5:
    print("Nota %d: %6.2f" % (x, notas[x]))
    x += 1
print("MÃ©dia: %5.2f" % (soma/len(notas)))


#Listagem 6.7 â€“ ApresentaÃ§Ã£o de nÃºmeros

nÃºmeros = [0,0,0,0,0]
x = 0 
while x < 5:
    nÃºmeros[x] = int(input("NÃºmero %d:" % (x+1))) 
    x += 1
while True:
    escolhido = int(input("Que posiÃ§Ã£o vocÃª quer imprimir (0 para sair): "))
    if escolhido == 0:
        break
    print("VocÃª escolheu o nÃºmero: %d" % (nÃºmeros[escolhido-1]))


#Listagem 6.8 â€“ Tentativa de copiar listas 

L = [1,2,3,4,5] 
V = L 
L 
V 
V[0] = 6 
V 
L 


#Listagem 6.9 â€“ CÃ³pia de listas 

L = [1,2,3,4,5] 
V = L[:] 
V[0] = 6 
L 
V 


#Listagem 6.10 â€“ Fatiamento de listas 

L = [1,2,3,4,5] 
L[0:5] 
L[:5] 
L[:-1] 
L[1:3] 
L[1:4] 
L[3:] 
L[:3] 
L[-1] 


#Listagem 6.11 â€“ Tamanho de listas 

L = [12,9,5] 
len(L) 
V=[] 
len(V) 


#Listagem 6.12 â€“ RepetiÃ§Ã£o com tamanho fixo da lista 

L = [1,2,3] 
x = 0 
while x < 3:
    print(L[x]) 
    x += 1 


#Listagem 6.13 â€“ RepetiÃ§Ã£o com tamanho da lista usando len() 

L = [1,2,3] 
x = 0 
while x < len(L):
    print(L[x]) 
    x += 1 


#Listagem 6.14 â€“ AdiÃ§Ã£o de elementos Ã  lista 

L = [] 
L.append("a") 
L 
L.append("b") 
L 
L.append("c") 
L 
len(L) 


#Listagem 6.15 â€“ AdiÃ§Ã£o de elementos Ã  lista 

L = [] 
while True:
    n = int(input("Digite um nÃºmero (0 sai):")) 	
    if n == 0: 	
        break 
    L.append(n) 
x = 0 
while x < len(L):	
    print(L[x]) 
    x = x+1


#Listagem 6.16 â€“ AdiÃ§Ã£o de listas 

L = [] 
L = L + [1] 
L 
L += [2] 
L 
L += [3,4,5] 
L 


#Listagem 6.17 â€“ AdiÃ§Ã£o de elementos e listas >>> L=["a"] 

L = ['a']
L.append("b") 
L 
L.extend(["c"]) 
L 
L.append(["d","e"]) 
L 
L.extend(["f","g","h"]) 
L 


#Listagem 6.18 â€“ AdiÃ§Ã£o de elementos e listas com append 

L = ["a"] 
L.append(["b"]) 
L.append(["c","d"]) 
len(L) 
L[1] 
L[2] 
len(L[2]) 
L[2][1]


#Listagem 6.19 â€“ RemoÃ§Ã£o de elementos 

L = ["a","b","c"] 
del L[1] 
L 
del L[0] 
L 


#Listagem 6.20 â€“ RemoÃ§Ã£o de fatias 

L = list(range(101)) 
del L[1:99] 
L 


#Listagem 6.21 â€“ SimulaÃ§Ã£o de uma fila de banco 

ultimo = 10 
fila = list(range(1,ultimo+1)) 
while True:
    print("\nExistem %d clientes na fila" % len(fila)) 
    print("Fila atual:", fila) 
    print("Digite F para adicionar um cliente ao fim da fila,") 
    print("ou A para realizar o atendimento. S para sair.") 
    operacao = input("Operação (F, A ou S):")
    if operacao == "A":
        if(len(fila))>0:
            atendido = fila.pop(0)
            print("Cliente %d atendido" % atendido) 
        else:
            print("Fila vazia! Ninguém para atender.") 
    elif operacao == "F":
        ultimo+=1 # Incrementa o ticket do novo cliente 
        fila.append(ultimo)
    elif operacao == "S":
            break
    else:
        print("Operção inválida! Digite apenas F, A ou S!")


#Listagem 6.22 â€“ Pilha de pratos 

prato = 5 
pilha = list(range(1,prato+1)) 
while True:
    print("\nExistem %d pratos na pilha" % len(pilha))
    print("Pilha atual:", pilha) 
    print("Digite E para empilhar um novo prato,") 
    print("ou D para desempilhar. S para sair.") 
    operacao = input("operacao (E, D ou S):") 
    if operacao == "D":
        if(len(pilha))>0:
            lavado = pilha.pop(-1) 
            print("Prato %d lavado" % lavado) 
        else: 
            print("Pilha vazia! Nada para lavar.") 
    elif operacao == "E":
        prato+=1 # Novo prato 
        pilha.append(prato) 
    elif operacao == "S": 
        break 
else:   
    print("operacao invÃ¡lida! Digite apenas E, D ou S!")


# Listagem 6.23 â€“ Pesquisa sequencial 

L = [15,7,27,39] 
p = int(input("Digite o valor a procurar:")) 
achou = False
x = 0 
while x < len(L):
    if L[x] == p:
        achou = True
        break
    x += 1 
if achou:
	print("%d achado na posiÃ§Ã£o %d" % (p,x)) 
else: 
	print("%d nÃ£o encontrado" % p)


#Listagem 6.44 â€“ OrdenaÃ§Ã£o pelo mÃ©todo de bolhas
L = [7, 3, 4, 12, 8]
fim = 5
while fim > 1:
    trocou = False
    x = 0
    while x < (fim-1):
        if L[x] > L[x + 1]:
            trocou-True
            temp = L[x]
            L[x] = L[x + 1]
            L[x + 1] = temp
        x += 1
    if not trocou:
        break
    fim -= 1
for e in L:
    print(e)



#Exemplo de estrutura match case no Python, equivalente ao switch em varias linguagens


opcao = int(input('Escolha a opÃ§Ã£o desejada: de 1 a 5: '))
 
#Estrutura Match Case

match opcao:
    
    case 1:
        print('Vc optou pela alternativa 1')
    
    case 2:
        print('Vc optou pela alternativa 2')

    case 3:
        print('Vc optou pela alternativa 3')
        
    case 4:
        print('Vc optou pela alternativa 4')

    case 5:
        print('Vc optou pela alternativa 5')
    
    case _:
        print('opÃ§Ã£o invÃ¡lida')

    
# Estrutura Match Case - Exercicio 1
'''
Crie um algoritmo e o respectivo cÃ³digo para mostrar ao usuÃ¡rio, primeiramente, 
um menu de opÃ§Ãµes e o numero da plataforma, do trem que ele deve tomar, 
conforme tabela abaixo, e uma mensagem de erro quando o destino nÃ£o existir:
Destino	NÃºmero da Plataforma
SÃ£o Paulo	Plataforma 1
Rio de Janeiro	Plataforma 2
Curitiba	Plataforma 3
Brasilia	Plataforma 4
FlorianÃ³polis	Plataforma 5
'''


print('Bem vindo a EstaÃ§Ã£o FIAP 1ESPI')
print('Escola o destino abaixo para selecionar o nÃºmero da plataforma:  ')
print('SÃ£o Paulo')
print('Rio de Janeiro')
print('Curitiba')
print('Brasilia')
print('FlorianÃ³polis')

destino = input('selecione a cidade de destino:   ')

match destino:
    case 'SÃ£o Paulo':
        print('Plataforma #1')
    
    case 'Rio de Janeiro':
        print('Plataforma #2')
    
    case 'Curitiba':
        print('Plataforma #3')
    
    case 'Brasilia':
        print('Plataforma #4')
    
    case 'FlorianÃ³polis':
        print('Plataforma #5')
    
    case _:
        print('NÃ£o existe o destino selecionado cabeÃ§a de pastel')
        
        
# Estrutura Match Case - Exercicio 2
'''
Crie um algoritmo e o respectivo cÃ³digo para mostrar ao usuÃ¡rio no restaurante, 
uma carta de vinho com 5 opÃ§Ãµes de cÃ³digo, para que ele faÃ§a sua escolha, 
e retorne o nome do vinho e o preÃ§o da garrafa 
1	Casa Valduga Gran	R$ 160,00
2	Casa Silva Gran Terrois los CarmÃ©nÃ¨re	R$ 130,00
3	Chocalan Origen Viognier	R$ 180,00
4	Miolo Demi-sec Reserva especial 2005	R$ 90,00
5	Casa Silva Terrois Cabernet Savignon	R$ 100,00
'''
print('Bem vindo ao restaturante da Vinheria Agnello!')
print('')
print('Selecione o cÃ³digo do vinho a ser consumido: ') 
print('- - - - - - - - - - - - - - - - - - - - - - -')
print('Casa Valduga Gran - Codigo 1')
print('Casa Silva Gran Terrois los CarmÃ©nÃ¨re -  cÃ³digo 2')
print('Chocalan Origen Viognier - cÃ³digo 3') 
print('Miolo Demi-sec Reserva especial 2005 - cÃ³digo 4')
print('Casa Silva Terrois Cabernet Savignon - cÃ³digo 5') 
print('- - - - - - - - - - - - - - - - - - - - - - -')
vinho = int(input('Selecione o vinho: '))      
print('- - - - - - - - - - - - - - - - - - - - - - -')
print('')
match vinho:
    
    case 1:
        print('vinho Casa Valduga Gran - R$ 160,00')
    
    case 2:
        print('Vinho Casa Silva Gran Terrois los CarmÃ©nÃ¨re - R$ 130,00 ')
    
    case 3:
        print('Vinho Chocalan Origen Viognier  - R$ 180,00')
    
    case 4:
        print('Miolo Demi-sec Reserva especial 2005 - R$ 90,00')
        
    case 5:
        print('Vinho Casa Silva Terrois Cabernet Savignon  - R$ 100,00')
    
    case _:
        print('a opÃ§Ã£o deve ser de 1 a 5')
   
        
# Calculadora de 4 operaÃ§Ãµes utilizando a estrutura match-case - Exercicio 3

p = float(input('Digite o primeiro nÃºmero: '))
s = float(input('Digite o segundo nÃºmero: '))

# Exibir o menu de opÃ§Ãµes:
print('1 - Soma')
print('2 - SubtraÃ§Ã£o')
print('3 - MultiplicaÃ§Ã£o')
print('4 - DivisÃ£o')
opcao = int(input('Selecione uma opÃ§Ã£o para a operacao matemÃ¡tica: '))

# Verificar qual a opÃ§Ã£o ecolhida:

match opcao:
    case 0:
        r = p + s
        print(f'A soma de {p} e {s} Ã© {r}')
    case 2:
        r = p - s
        print(f'A subtraÃ§Ã£o de {p} e {s} Ã© {r}')
    case 3:
        r = p * s
        print(f'A multiplicaÃ§Ã£o de {p} e {s} Ã© {r}')
    case 4:
        if s != 0:
            r = p / s
            print(f'A divisÃ£o de {p} e {s} Ã© {r}')
        else:
            print('NÃ£o Ã© possÃ­vel realizar uma divisÃ£o por zero')
    case _:                 # default
        print('OpÃ§Ã£o InvÃ¡lida')


# Classificar se a letra Ã© vogal ou consoante - Exercicio #4

letra = input('Informe uma letra: ')
match letra:
    case 'a' | 'e' | 'i' | '0' | 'u' | 'A' | 'E' | 'I' | 'O' | 'U':
        print('Vogal')
    case _:
        print('Consoante')

# Mostrar todos os primos atÃ© um numero N dado e fazer uma contagem 


n = int(input('Digite um nÃºmero inteiro positivo: '))

if n <= 1:
    print(n, 'nÃ£o Ã© primo')
else:
    primos = []
    for i in range(2, n+1):
        primo = True
        for j in range(2, i):
            if i % j == 0:
                primo = False
                break
        if primo:
            primos.append(i)
    
    print('NÃºmeros primos atÃ© ', n, ':')
    print(primos)
    print('Total de nÃºmeros primos:', len(primos))


# Exemplo de tratamento de exceÃ§Ã£o - Try - Except - Else - finally

num1 = 10
num2 = 0

try:
    resultado = num1 / num2

except ZeroDivisionError:
    print("Erro: divisÃ£o por zero")

else:
    print(resultado)

finally:
    print("Fim do programa - Entendi!!!!!!")

# Exemplo de definicao de função no Python

def soma(a, b):
    resultado = a + b
    return resultado

s = soma (1,2)
print(s)

# Outro exemplo de definicao de função no Python

10


# Outro exemplo ainda de definicao de função no Python sem colocar parÃ¢metros

print(' -'*30)
print('          Vinheria Agnello')
print(' -'*30)
print('')
print(' -'*30)
print('          PromoÃ§Ã£o de vinhos - queima de estoque')
print(' -'*30)


def linha_separacao():
    return print(' -'*30)

linha_separacao()

print('          PromoÃ§Ã£o de vinhos - queima de estoque')    

linha_separacao()

print('          PromoÃ§Ã£o de vinhos - Queima de Estoque')
linha_separacao()





'''
SEGUNDO SEMESTRE
EXERCICIOS DE FUNÃ‡ÃƒO
'''

#importacao de bibliotecas
import math
import Numpy as np
import Pandas as pd
import matplotlib as plt




# Exemplo de definicao de função no Python

def soma(a, b):
    resultado = a + b
    return resultado

s = soma (1,2)
print(s)


# Outro exemplo de definicao de função no Python


# Outro exemplo ainda de definicao de função no Python sem colocar parÃ¢metros

print(' -'*30)
print('          Vinheria Agnello')
print(' -'*30)
print('')
print(' -'*30)
print('          PromoÃ§Ã£o de vinhos - queima de estoque')
print(' -'*30)


def linha_separacao():
    return print(' -'*30)

linha_separacao()

print('          PromoÃ§Ã£o de vinhos - queima de estoque')    

linha_separacao()

print('          PromoÃ§Ã£o de vinhos - Queima de Estoque')
linha_separacao()



# Exercicio 1 - Crie uma função para calcular o dobro de um numero


# DefiniÃ§Ã£o de função

def dobrar_numero():
    numero = float(input("Digite um nÃºmero: "))
    resultado = numero * 2
    return resultado

print(' O dobro do numero Ã©:   ',dobrar_numero())



#  Exercicio 2 - Crie uma função que soma dois numeros inteiros 

# DefiniÃ§Ã£o de função
def somar_numeros():
    numero1 = float(input("Digite o primeiro nÃºmero: "))
    numero2 = float(input("Digite o segundo nÃºmero: "))
    resultado = numero1 + numero2
    return resultado

print(' A soma dos numeros Ã©:   ',somar_numeros())



#  Exercicio 3 - Crie uma função que forneÃ§a o mÃ³dulo do numero

# DefiniÃ§Ã£o de função

def valor_absoluto():
    numero = float(input("Digite um nÃºmero:  "))
    resultado = abs(numero)
    return resultado

print(' O modulo do numero Ã©:  ',valor_absoluto())



#  Exercicio 4 - Crie uma função que forneÃ§a o quadrado do numero

# DefiniÃ§Ã£o de função

def quadrado_de_n():
    numero = float(input("Digite um nÃºmero: "))
    resultado = numero ** 2
    return resultado

print(' O modulo do numero Ã©:  ',quadrado_de_n())



 #  Exercicio 5 - Crie um cÃ³digo para calcular o Fatorial de um numero

# DefiniÃ§Ã£o de função
def fatorial(n):
    if n == 0 or n == 1:
        return 1
    else:
        return n * fatorial(n-1)

n = int(input('Digite o numero que vc quer calcular o fatorial:  '))

print("O fatorial de", n, "Ã©", fatorial(n))


#  Exercicio 6 - Crie uma função que retorna o maior entre dois nÃºmeros:

# DefiniÃ§Ã£o de função
    
def busca_maior():
    numero1 = float(input("Digite o primeiro nÃºmero: "))
    numero2 = float(input("Digite o segundo nÃºmero: "))
    if numero1 > numero2:
        return numero1
    else:
        return numero2

print(' O maior numero Ã©:  ',busca_maior())



#  Exercicio 7 - Crie uma função que retorne o antecessor do numero

# DefiniÃ§Ã£o de função
def buscar_antecessor():
    numero = float(input("Digite um nÃºmero: "))
    antecessor = numero - 1
    return antecessor

print(' O antecessor Ã©:  ',buscar_antecessor())



#  Exercicio 8 - Crie uma função que retorna o menor entre dois nÃºmeros

# DefiniÃ§Ã£o doe função

def buscar_menor():
    numero1 = float(input("Digite o primeiro nÃºmero: "))
    numero2 = float(input("Digite o segundo nÃºmero: "))
    if numero1 < numero2:
        return numero1
    else:
        return numero2

print('o numero menor Ã©: ', buscar_menor())



#  Exercicio 9 - Crie uma função que verifica se um nÃºmero Ã© par:

# DefiniÃ§Ã£o de função

def verificar_par():
    numero = int(input("Digite um nÃºmero: "))
    if numero % 2 == 0:
        return print('O numero Ã© par!')
    else:
        return print('O numero Ã© impar!')

print(verificar_par())



#  Exercicio 10 - Crie uma função para calcular a raiz quadrada de um numero:

import math

    
# DefiniÃ§Ã£o de função


def calcular_raiz_quadrada():
    numero = float(input("Digite um nÃºmero: "))
    raiz_quadrada = math.sqrt(numero)
    return raiz_quadrada

print('A raiz quadrada Ã©: ',calcular_raiz_quadrada())
      

# Exercicio 11 - Crie um codigo para somar dois numeros complexos

# DefiniÃ§Ã£o de função

def somar_numeros_complexos(num1, num2):
    real = num1[0] + num2[0]
    imaginaria = num1[1] + num2[1]
    resultado = (real, imaginaria)
    return resultado

# Solicita as partes real e imaginÃ¡ria do primeiro nÃºmero complexo ao usuÃ¡rio
real1 = float(input("Digite a parte real do primeiro nÃºmero complexo: "))
imaginaria1 = float(input("Digite a parte imaginÃ¡ria do primeiro nÃºmero complexo: "))

# Solicita as partes real e imaginÃ¡ria do segundo nÃºmero complexo ao usuÃ¡rio
real2 = float(input("Digite a parte real do segundo nÃºmero complexo: "))
imaginaria2 = float(input("Digite a parte imaginÃ¡ria do segundo nÃºmero complexo: "))

# Cria as representaÃ§Ãµes dos nÃºmeros complexos como tuplas (parte real, parte imaginÃ¡ria)
num1 = (real1, imaginaria1)
num2 = (real2, imaginaria2)

# Realiza a soma dos nÃºmeros complexos
soma = somar_numeros_complexos(num1, num2)

# Exibe o resultado
print("A soma dos nÃºmeros complexos Ã©:", soma)



'''
 Exercicio 12 - Crie um codigo completo e detalhado que retorne as duas raizes de 
 uma equaÃ§Ã£o de segundo grau usando a formula de bascara
'''

import math

# DefiniÃ§ao de função

def calcular_raizes(a, b, c):
    discriminante = b**2 - 4*a*c

    if discriminante > 0:
        raiz1 = (-b + math.sqrt(discriminante)) / (2*a)
        raiz2 = (-b - math.sqrt(discriminante)) / (2*a)
        return raiz1, raiz2

    elif discriminante == 0:
        raiz = -b / (2*a)
        return raiz, raiz

    else:
        parte_real = -b / (2*a)
        parte_imaginaria = math.sqrt(-discriminante) / (2*a)
        raiz1 = complex(parte_real, parte_imaginaria)
        raiz2 = complex(parte_real, -parte_imaginaria)
        return raiz1, raiz2

# Solicita os coeficientes da equaÃ§Ã£o de segundo grau ao usuÃ¡rio
coeficiente_a = float(input("Digite o coeficiente a: "))
coeficiente_b = float(input("Digite o coeficiente b: "))
coeficiente_c = float(input("Digite o coeficiente c: "))

# Calcula as raÃ­zes
raiz1, raiz2 = calcular_raizes(coeficiente_a, coeficiente_b, coeficiente_c)

# Exibe as raÃ­zes
print("As raÃ­zes da equaÃ§Ã£o sÃ£o:")
print("Raiz 1:", raiz1)
print("Raiz 2:", raiz2)



#  Exercicio 13 - Crie um codigo para calcular o mudulo de um vetor... em N dimensÃµes

import math

#definiÃ§Ã£o de função

def calcular_modulo_vetor(vetor):
    soma_quadrados = 0
    for componente in vetor:
        soma_quadrados += componente**2
    modulo = math.sqrt(soma_quadrados)
    return modulo

# Solicita as componentes do vetor ao usuÃ¡rio
dimensoes = int(input("Digite o nÃºmero de dimensÃµes do vetor: "))
vetor = []
for i in range(dimensoes):
    componente = float(input(f"Digite a componente {i+1} do vetor: "))
    vetor.append(componente)

# Calcula o mÃ³dulo do vetor
modulo = calcular_modulo_vetor(vetor)

# Exibe o resultado
print("O mÃ³dulo do vetor Ã©:", modulo)

'''
Exercicio 14 - Crie um codigo utilizando função para criar a sequencia de Fibonacci para os 
primeiro n elementos. Apresente-os e a seguir informe quais desses sÃ£o numeros primo

'''

# importa modulo sys
import sys

# DefiniÃ§Ã£o de função

 # Inicializa a sequÃªncia de Fibonacci com os dois primoiros nÃºmeros
def fibonacci(n):
    Sequencia_Fibonacci = [1, 1] 

    while len(Sequencia_Fibonacci) < n:
        proximo_numero = Sequencia_Fibonacci[-1] + Sequencia_Fibonacci[-2]  # Calcula o prÃ³ximo nÃºmero da sequÃªncia
        Sequencia_Fibonacci.append(proximo_numero)  # Adiciona o prÃ³ximo nÃºmero Ã  lista da sequÃªncia

    return Sequencia_Fibonacci


def eh_primo(numero):
    if numero <= 1:
        return False
    elif numero == 2:
        return True

    for i in range(2, int(numero ** 0.5) + 1):
        if numero % i == 0:
            return False

    return True

n = int(input("Digite o valor de 'n': "))
fibonacci_seq = fibonacci(n)

# Mostra a sequÃªncia de Fibonacci
print("SequÃªncia de Fibonacci:")
print(fibonacci_seq)

# Mostra quais nÃºmeros sÃ£o primos
primos = [numero for numero in fibonacci_seq if eh_primo(numero)]
print("NÃºmeros primos na sequÃªncia de Fibonacci:")
print(primos)



# Exercicio 15 - entendendo parÃ¢metros, argumentos, *args e **kwargs

def calculadora_salario(valor, horas=220):
    return horas * valor

valor_total = calculadora_salario(299.25)

print(valor_total)

'''
Exercicios envolvendo args e kwargs
'''

'''
exercicio 15 - Crie uma função chamada soma que recebe argumentos posicionais 
variÃ¡veis e retorna a soma de todos os nÃºmeros passados como argumentos.
'''

#DefiniÃ§Ã£o de função

def soma(*args):
    return sum(args)

resultado = soma(2, 4, 6, 8)
print(resultado)  # Deve imprimir 20


'''
ExercÃ­cio 16: Calculadora AvanÃ§ada

Crie uma função chamada calculadora que aceita argumentos nomeados 
variÃ¡veis para realizar operaÃ§Ãµes matemÃ¡ticas bÃ¡sicas. 
Os argumentos possÃ­veis sÃ£o: num1, num2, e operacao. 
As operaÃ§Ãµes possÃ­veis sÃ£o: "soma", "subtracao", "multiplicacao" e "divisao".
'''

#DefiniÃ§Ã£o de função

def calculadora(**kwargs):
    num1 = kwargs.get('num1', 0)
    num2 = kwargs.get('num2', 0)
    operacao = kwargs.get('operacao', 'soma')

    if operacao == 'soma':
        return num1 + num2
    elif operacao == 'subtracao':
        return num1 - num2
    elif operacao == 'multiplicacao':
        return num1 * num2
    elif operacao == 'divisao':
        if num2 != 0:
            return num1 / num2
        else:
            return "Erro: divisÃ£o por zero"

resultado = calculadora(num1=10, num2=5, operacao='subtracao')
print(resultado)  # Deve imprimir 5


'''
ExercÃ­cio 17: MÃ©dia simples

Crie uma função que recebe um nÃºmero indeterminado de argumentos nomeados 
representando as notas 
'''

#Definicao de função

def calcula_media(*args):
    return sum(args) / len(args)

numeros = [float(x) for x in input("Digite uma lista de nÃºmeros separados por espaÃ§o: ").split()]
resultado = calcula_media(*numeros)
print(f"A mÃ©dia Ã©: {resultado}")


'''
ExercÃ­cio 18:  ConcatenaÃ§Ã£o de Strings

Crie uma função que recebe strings para serem concatenadas
'''

#Definicao de função

def concatena_strings(*args):
    return ''.join(args)

strings = input("Digite uma lista de strings separadas por espaÃ§o: ").split()
resultado = concatena_strings(*strings)
print(f"A concatenaÃ§Ã£o das strings Ã©: {resultado}")

'''
ExercÃ­cio 19:  Contagem de Vogais

Crie uma função que faÃ§a contagem de vogais.
lembre-se que o Python Ã© Case sensitive
'''

#Definicao de função

def conta_vogais(*args):
    vogais = 'aeiouAEIOU'
    count = 0
    for arg in args:
        count += sum(1 for char in arg if char in vogais)
    return count

strings = input("Digite uma lista de strings separadas por espaÃ§o: ").split()
resultado = conta_vogais(*strings)
print(f"Total de vogais: {resultado}")


'''
ExercÃ­cio 20: Calculadora de Desconto

Crie uma função para criar uma calculadora para calculo de desconto

'''

#Definicao de função

def calcula_desconto(**kwargs):
    total = kwargs.get('total', 0)
    desconto = kwargs.get('desconto', 0.1)
    valor_com_desconto = total * (1 - desconto)
    return valor_com_desconto

total = float(input("Digite o valor total: "))
desconto = float(input("Digite a taxa de desconto (em decimal): "))
resultado = calcula_desconto(total=total, desconto=desconto)
print(f"Total com desconto: {resultado}")


'''
ExercÃ­cio 21: OrdenaÃ§Ã£o de NÃºmeros

Crie uma função que ordene uma lista de numeros

'''

#Definicao de função

def ordena_numeros(*args):
    return sorted(args)

numeros = [float(x) for x in input("Digite uma lista de nÃºmeros separados por espaÃ§o: ").split()]
resultado = ordena_numeros(*numeros)
print(f"NÃºmeros ordenados: {resultado}")


'''
ExercÃ­cio 22: ConversÃ£o de Unidades

Crie uma função que converta de centimentos para metros e de metros
para centimentos respectivamente

'''

#Definicao de função

def converte_unidades(**kwargs):
    valor = kwargs.get('valor', 0)
    unidade = kwargs.get('unidade', 'cm')

    if unidade == 'cm':
        return f"{valor} cm Ã© igual a {valor * 0.01} m"
    elif unidade == 'm':
        return f"{valor} m Ã© igual a {valor * 100} cm"

valor = float(input("Digite o valor: "))
unidade = input("Digite a unidade (cm/m): ")
resultado = converte_unidades(valor=valor, unidade=unidade)
print(resultado)


'''
ExercÃ­cio 23:  Lista de Compras

Crie uma função que crie uma  Lista de Compras

'''

#Definicao de função

def lista_de_compras(**kwargs):
    itens = kwargs.get('itens', [])
    return itens

itens = input("Digite uma lista de itens de compra separados por espaÃ§o: ").split()
compras = lista_de_compras(itens=itens)
print(f"Lista de compras: {compras}")


'''
ExercÃ­cio 24:   Contagem de args

Escreva uma função que conte e retorne quantos argumentos foram passados para ela.

'''

#Definicao de função

def contar_args(*args):
    return len(args)

numeros = input("Digite nÃºmeros separados por espaÃ§o: ").split()
resultado = contar_args(*numeros)
print("Total de argumentos:", resultado)


'''
ExercÃ­cio 25:   função com kwargs para informaÃ§Ãµes de usuÃ¡rio

Crie uma função que receba informaÃ§Ãµes de usuÃ¡rio (nome, idade, cidade) 
como argumentos nomeados e imprima-as.

'''

#Definicao de função

def info_usuario(**kwargs):
    print("Nome:", kwargs.get('nome'))
    print("Idade:", kwargs.get('idade'))
    print("Cidade:", kwargs.get('cidade'))

info = {
    'nome': input("Nome: "),
    'idade': input("Idade: "),
    'cidade': input("Cidade: ")
}
info_usuario(**info)

'''
ExercÃ­cio 26:   função com *args e kwargs para saudaÃ§Ãµes
Crie uma função que aceite o nome do usuÃ¡rio como argumento posicional (*args) 
e outras informaÃ§Ãµes como argumentos nomeados (**kwargs) para gerar
saudaÃ§Ãµes personalizadas.

'''

#Definicao de função

def saudacao_personalizada(*args, **kwargs):
    nome = args[0]
    saudacao = kwargs.get('saudacao', 'OlÃ¡')
    sobrenome = kwargs.get('sobrenome', '')
    idade = kwargs.get('idade', '')
    print(f"{saudacao}, {nome} {sobrenome}! {idade}")

nome = input("Nome: ")
sobrenome = input("Sobrenome: ")
idade = input("Idade: ")
saudacao_personalizada(nome, saudacao="Oi", sobrenome=sobrenome, idade=idade)


'''
Exercicio 27: CÃ¡lculos com kwargs
Crie uma função que aceite argumentos nomeados (**kwargs) representando 
valores e operaÃ§Ãµes matemÃ¡ticas, e retorne o resultado dessas operaÃ§Ãµes.
'''

#DefiniÃ§Ã£o de função

def calcular(**kwargs):
    valor = kwargs['valor']
    operacoes = kwargs['operacoes']
    resultado = valor
    for operacao in operacoes:
        if operacao[0] == '+':
            resultado += operacao[1]
        elif operacao[0] == '-':
            resultado -= operacao[1]
        elif operacao[0] == '*':
            resultado *= operacao[1]
        elif operacao[0] == '/':
            resultado /= operacao[1]
    return resultado

valor = float(input("Digite um valor: "))
operacoes = []
while True:
    op = input("Digite a operacao (+, -, *, /) e o valor separados por espaÃ§o (ou 'fim' para parar): ")
    if op == 'fim':
        break
    operacao, valor_op = op.split()
    operacoes.append((operacao, float(valor_op)))

resultado = calcular(valor=valor, operacoes=operacoes)
print("Resultado:", resultado)

'''
Exercicio 28: CÃ¡lculos com kwargs
função com kwargs para converter moedas
Crie uma função que aceite um valor em dÃ³lares e uma taxa de conversÃ£o como 
argumentos nomeados (**kwargs) e retorne o valor convertido para a moeda especificada.
'''

#DefiniÃ§Ã£o de função

def converter_moeda(**kwargs):
    valor_dolares = kwargs.get('valor_dolares', 0)
    taxa = kwargs.get('taxa', 1)
    valor_convertido = valor_dolares * taxa
    return valor_convertido

valor = float(input("Digite o valor em dÃ³lares: "))
taxa = float(input("Digite a taxa de conversÃ£o: "))
valor_convertido = converter_moeda(valor_dolares=valor, taxa=taxa)
print(f"Valor convertido: {valor_convertido:.2f}")


'''
Exercicio 29: função com args para verificar nÃºmeros primos
Crie uma função que aceite nÃºmeros como argumentos *args e retorne 
uma lista contendo os nÃºmeros primos dessa lista.

'''

#DefiniÃ§Ã£o de funções

def eh_primo(num):
    if num <= 1:
        return False
    for i in range(2, int(num ** 0.5) + 1):
        if num % i == 0:
            return False
    return True

def encontrar_primos(*args):
    primos = [num for num in args if eh_primo(num)]
    return primos

numeros = [int(x) for x in input("Digite nÃºmeros separados por espaÃ§o: ").split()]
primos_encontrados = encontrar_primos(*numeros)
print("NÃºmeros primos:", primos_encontrados)


'''
Exercicio 30: função com args para verificar nÃºmeros primos
Crie uma função que aceite nÃºmeros como argumentos *args e retorne 
uma lista contendo os nÃºmeros primos dessa lista.

'''

#DefiniÃ§Ã£o de funções

def calcular_juros_compostos(**kwargs):
    principal = kwargs.get('principal', 0)
    taxa_juros = kwargs.get('taxa_juros', 0)
    periodo = kwargs.get('periodo', 0)
    montante = principal * (1 + (taxa_juros / 100)) ** periodo
    return montante

principal = float(input("Digite o valor principal: "))
taxa = float(input("Digite a taxa de juros (em %): "))
periodo = int(input("Digite o perÃ­odo (em anos): "))
montante_final = calcular_juros_compostos(principal=principal, taxa_juros=taxa, periodo=periodo)
print(f"Montante final: R$ {montante_final:.2f}")

'''
EXERCICIOS ENVOLVENDO DICIONARIO
'''

'''
Exercicio 31 - Enunciado: Crie um dicionÃ¡rio vazio e imprima-o
'''

dictio = {}
print(dictio)

'''
ExercÃ­cio 32 - Enunciado: Crie um dicionÃ¡rio com trÃªs pares chave-valor 
representando nomes de frutas e suas quantidades. Imprima o dicionÃ¡rio.
'''

frutas = {'maca': 5, 'banana': 3, 'laranja': 7}
print(frutas)

'''
ExercÃ­cio 33 - Enunciado: Adicione uma nova fruta ao dicionÃ¡rio do exercÃ­cio 32
com sua respectiva quantidade. Imprima o dicionÃ¡rio atualizado.
'''

frutas['uva'] = 9
print(frutas)

'''
ExercÃ­cio 34 - Enunciado: Dado um dicionÃ¡rio de notas dos alunos, 
calcule a mÃ©dia das notas e imprima o resultado.
'''

notas = {'Alice': 85, 'Bob': 92, 'Carol': 78}
media = sum(notas.values()) / len(notas)
print(f"A mÃ©dia das notas Ã©: {media:.2f}")

'''
ExercÃ­cio 35 - Enunciado: Crie um dicionÃ¡rio de palavras e suas traduÃ§Ãµes 
em outro idioma. PeÃ§a ao usuÃ¡rio para digitar uma palavra e retorne 
a traduÃ§Ã£o correspondente, se existir.
'''

translations = {'hello': 'olÃ¡', 'dog': 'cÃ£o', 'cat': 'gato'}
palavra = input("Digite uma palavra: ")
translation = translations.get(palavra, "TraduÃ§Ã£o nÃ£o encontrada")
print(translation)

'''
ExercÃ­cio 36 - Enunciado: Crie um dicionÃ¡rio com nomes de cidades 
como keys e suas populaÃ§Ãµes como valores. PeÃ§a ao usuÃ¡rio para 
digitar uma cidade e mostre a populaÃ§Ã£o correspondente.
'''

cidades = {'New York': 8623000, 'Los Angeles': 3990000, 'Chicago': 2716000}
cidade = input("Digite o nome da cidade: ")
populacao = cidades.get(cidade, "Cidade nÃ£o encontrada")
print(f"A populaÃ§Ã£o de {cidade} Ã© {populacao}")

'''
ExercÃ­cio 37 - Enunciado: Crie um dicionÃ¡rio com nomes de produtos e
seus preÃ§os. Calcule o valor total de uma lista de compras fornecida pelo usuÃ¡rio.
'''

produtos = {'maca': 1.25, 'banana': 0.75, 'laranja': 0.90}
shopping_list = input("Digite os produtos da lista (separados por vÃ­rgula): ").split(',')
total_cost = sum(produtos[item] for item in shopping_list if item in produtos)
print(f"O custo total Ã©: R${total_cost:.2f}")


'''
ExercÃ­cio 38 - Enunciado: Crie um programa que simule um dicionÃ¡rio de contatos. 
Permita ao usuÃ¡rio adicionar, remover e buscar contatos pelo nome.
'''

contatos = {}

while True:
    print("1. Adicionar contato")
    print("2. Remover contato")
    print("3. Buscar contato")
    print("4. Sair")
    choice = int(input("choice uma opÃ§Ã£o: "))
    
    if choice == 1:
        nome = input("Digite o nome do contato: ")
        numero = input("Digite o nÃºmero do contato: ")
        contatos[nome] = numero
    elif choice == 2:
        nome = input("Digite o nome do contato a ser removido: ")
        if nome in contatos:
            del contatos[nome]
            print(f"Contato {nome} removido.")
        else:
            print("Contato nÃ£o encontrado.")
    elif choice == 3:
        nome = input("Digite o nome do contato a ser buscado: ")
        numero = contatos.get(nome, "Contato nÃ£o encontrado")
        print(f"NÃºmero de {nome}: {numero}")
    elif choice == 4:
        break
    
'''    
ExercÃ­cio 39 - Enunciado: Crie um dicionÃ¡rio contendo nomes de paÃ­ses e 
suas capitais. PeÃ§a ao usuÃ¡rio para adivinhar a capital de um paÃ­s escolhido aleatoriamente.
'''

import random

paises = {'Brasil': 'BrasÃ­lia', 'EUA': 'Washington', 'FranÃ§a': 'Paris'}
pais = random.choice(list(paises.keys()))
guess = input(f"Qual Ã© a capital de {pais}? ")
if guess == paises[pais]:
    print("Resposta correta!")
else:
    print(f"Resposta incorreta. A capital de {pais} Ã© {paises[pais]}.")

'''
ExercÃ­cio 40 - Enunciado: Crie um dicionÃ¡rio com palavras-chave e suas 
definiÃ§Ãµes. Desafie o usuÃ¡rio a adivinhar uma definiÃ§Ã£o 
a partir de uma palavra-chave.
'''

definicoes = {'python': 'Linguagem de programaÃ§Ã£o de alto nÃ­vel', 'database': 'Conjunto organizado de dados'}
palavra_chave = random.choice(list(definicoes.keys()))
definicao = input(f"Qual Ã© a definiÃ§Ã£o de '{palavra_chave}'? ")
if definicao == definicoes[palavra_chave]:
    print("Resposta correta!")
else:
    print(f"Resposta incorreta. A definiÃ§Ã£o de '{palavra_chave}' Ã© '{definicoes[palavra_chave]}'.")

'''
ExercÃ­cio 41 - Enunciado: Crie um programa de cadastro de funcionÃ¡rios 
utilizando dicionÃ¡rios para armazenar informaÃ§Ãµes como nome, cargo e salÃ¡rio.
'''

empregados = []
while True:
    print("1. Cadastrar funcionÃ¡rio")
    print("2. Listar funcionÃ¡rios")
    print("3. Sair")
    choice = int(input("choice uma opÃ§Ã£o: "))
    
    if choice == 1:
        empregado = {}
        empregado['nome'] = input("Nome: ")
        empregado['cargo'] = input("Cargo: ")
        empregado['salario'] = float(input("SalÃ¡rio: "))
        empregados.append(empregado)
        print("FuncionÃ¡rio cadastrado.")
    elif choice == 2:
        for empregado in empregados:
            print(f"Nome: {empregado['nome']}, Cargo: {empregado['cargo']}, SalÃ¡rio: {empregado['salario']:.2f}")
    elif choice == 3:
        break

'''
ExercÃ­cio 42 - Enunciado: Crie um dicionÃ¡rio com nomes de filmes e seus anos
de lanÃ§amento. PeÃ§a ao usuÃ¡rio para digitar um ano e exiba os filmes lanÃ§ados naquele ano.
'''

filmes = {'Matrix': 1999, 'Senhor dos AnÃ©is': 2001, 'Avatar': 2009}
ano = int(input("Digite um ano: "))
filmes_no_ano = [movie for movie, release_ano in filmes.items() if release_ano == ano]
if filmes_no_ano:
    print(f"Filmes lanÃ§ados em {ano}: {', '.join(filmes_no_ano)}")
else:
    print(f"Nenhum filme lanÃ§ado em {ano}.")
'''
ExercÃ­cio 43 - Enunciado: Crie um dicionÃ¡rio representando uma agenda de compromissos. 
Permita ao usuÃ¡rio adicionar e listar compromissos para um determinado dia.
'''

compromissos = {}
while True:
    print("1. Adicionar compromisso")
    print("2. Listar compromissos")
    print("3. Sair")
    choice = int(input("escolha uma opÃ§Ã£o: "))
    
    if choice == 1:
        date = input("Digite a data (DD/MM/AAAA): ")
        compromisso = input("Digite o compromisso: ")
        if date in compromissos:
            compromissos[date].append(compromisso)
        else:
            compromissos[date] = [compromisso]
        print("Compromisso adicionado.")
    elif choice == 2:
        date = input("Digite a data para istar os compromissos: ")
        compromissos_list = compromissos.get(date, [])
        if compromissos_list:
            print(f"Compromissos em {date}:")
            for idx, compromisso in enumerate(compromissos_list, start=1):
                print(f"{idx}. {compromisso}")
        else:
            print(f"Nenhum compromisso marcado para {date}.")
    elif choice == 3:
        break

'''
ExercÃ­cio 44 - Enunciado: Crie um dicionÃ¡rio com nomes de animais como keys 
e suas caracterÃ­sticas como valores. 
PeÃ§a ao usuÃ¡rio para adivinhar um animal a partir de uma caracterÃ­stica.
'''


animais = {'cachorro': 'late', 'gato': 'mia', 'elefante': 'trumpeta'}
caracteristica = input("Digite uma caracterÃ­stica: ")
matching_animais = [animal for animal, char in animais.items() if char == caracteristica]
if matching_animais:
    print(f"Animais que {caracteristica}: {', '.join(matching_animais)}")
else:
    print(f"Nenhum animal encontrado com a caracterÃ­stica '{caracteristica}'.")

'''
ExercÃ­cio 45 - Enunciado: Crie um dicionÃ¡rio com nomes de pratos de 
diferentes paÃ­ses e suas receitas. 
Permita ao usuÃ¡rio escolher um prato e exibir a receita.
'''

pratos = {'feijoada': 'Receita da feijoada...', 'sushi': 'Receita de sushi...', 'pasta': 'Receita de pasta...'}
prato = input("Digite o nome de um prato: ")
receita = pratos.get(prato, "Prato nÃ£o encontrado")
print(f"Receita de {prato}: {receita}")


'''
ExercÃ­cio 46: - Enunciado: Crie um programa de quiz com perguntas sobre 
capitais de paÃ­ses. 
Use um dicionÃ¡rio para armazenar as perguntas e respostas corretas.
'''

questoes = {'Qual Ã© a capital do Brasil?': 'BrasÃ­lia', 'Qual Ã© a capital dos Estados Unidos?': 'Washington'}
score = 0
for questao, correct_answer in questoes.items():
    user_answer = input(questao + " ")
    if user_answer == correct_answer:
        print("Resposta correta!")
        score += 1
    else:
        print(f"Resposta incorreta. A resposta correta Ã©: {correct_answer}")
print(f"VocÃª acertou {score}/{len(questoes)} perguntas.")


'''
ExercÃ­cio 47 - Enunciado: Crie um dicionÃ¡rio com palavras-chave relacionadas a
um tÃ³pico de interesse. Implemente um sistema de busca para exibir 
informaÃ§Ãµes sobre a palavra-chave inserida pelo usuÃ¡rio.
'''

palavra_keys = {'python': 'Linguagem de programaÃ§Ã£o', 'machine learning': 'TÃ©cnica de aprendizado de mÃ¡quina', 'web development': 'Desenvolvimento de sites'}
pesquisar_termo = input("Digite uma palavra-chave: ")
definicao = palavra_keys.get(pesquisar_termo, "Palavra-chave nÃ£o encontrada")
print(f"{pesquisar_termo}: {definicao}")

'''
ExercÃ­cio 48 - Enunciado: Crie um jogo de adivinhaÃ§Ã£o onde o programa escolhe 
uma palavra e o jogador precisa adivinhar as letras. 
Use um dicionÃ¡rio para controlar as letras corretas e incorretas.
'''

import random

palavras = ['python', 'programming', 'dictionary', 'challenge']
palavra = random.choice(palavras)
correct_letters = set()
letras_incorretas = set()

while len(letras_incorretas) < 6:
    display_palavra = ''.join([letter if letter in correct_letters else '_' for letter in palavra])
    print(f"Palavra: {display_palavra}")
    print(f"Letras incorretas: {', '.join(letras_incorretas)}")
    
    guess = input("Digite uma letra: ")
    if guess in correct_letters or guess in letras_incorretas:
        print("VocÃª jÃ¡ tentou essa letra.")
    elif guess in palavra:
        correct_letters.add(guess)
        if all(letter in correct_letters for letter in palavra):
            print(f"ParabÃ©ns! A palavra era '{palavra}'.")
            break
    else:
        letras_incorretas.add(guess)
else:
    print(f"VocÃª errou muito! A palavra era '{palavra}'.")

'''
ExercÃ­cio 49: - Enunciado: Crie um dicionÃ¡rio com nomes de ingredientees 
e suas quantidades em gramas. PeÃ§a ao usuÃ¡rio para digitar o nome de um ingredientee
e a quantidade desejada, e ajuste o dicionÃ¡rio.
'''

ingredientes = {'farinha': 500, 'aÃ§Ãºcar': 300, 'ovos': 4}
ingrediente = input("Digite o nome do ingredientee: ")
quantidade = int(input("Digite a quantidade desejada (em gramas): "))
ingredientes[ingrediente] = quantidade
print(f"ingrediente '{ingrediente}' atualizado para {quantidade}g.")


'''
ExercÃ­cio 50 - Enunciado: Crie um programa que simule um carrinho de compras. 
Use um dicionÃ¡rio para armazenar os itens e suas quantidades. 
Permita ao usuÃ¡rio adicionar, remover e listar itens do carrinho.
'''

carrinho_de_compras = {}
while True:
    print("1. Adicionar item ao carrinho")
    print("2. Remover item do carrinho")
    print("3. Listar itens do carrinho")
    print("4. Sair")
    choice = int(input("choice uma opÃ§Ã£o: "))
    
    if choice == 1:
        item = input("Digite o nome do item: ")
        quantidade = int(input("Digite a quantidade desejada: "))
        if item in carrinho_de_compras:
            carrinho_de_compras[item] += quantidade
        else:
            carrinho_de_compras[item] = quantidade
        print(f"{quantidade} {item}(s) adicionado(s) ao carrinho.")
    elif choice == 2:
        item = input("Digite o nome do item a ser removido: ")
        if item in carrinho_de_compras:
            del carrinho_de_compras[item]
            print(f"{item} removido do carrinho.")
        else:
            print(f"{item} nÃ£o encontrado no carrinho.")
    elif choice == 3:
        print("Itens no carrinho:")
        for item, quantidade in carrinho_de_compras.items():
            print(f"{item}: {quantidade}")
    elif choice == 4:
        break

'''
ExercÃ­cio 51 - Gerenciamento de Estoque
Enunciado - VocÃª Ã© responsÃ¡vel por desenvolver um sistema de gerenciamento de estoque 
para uma loja. O sistema deve permitir a adiÃ§Ã£o de novos produtos, 
atualizaÃ§Ã£o de quantidades em estoque, venda de produtos e exibiÃ§Ã£o de relatÃ³rios. 
VocÃª deve usar um dicionÃ¡rio para armazenar as informaÃ§Ãµes dos produtos.
 - Implemente as seguintes funcionalidades:
*1* - Adicionar Produto: O usuÃ¡rio pode adicionar um novo produto ao estoque. 
      Cada produto terÃ¡ um nome, preÃ§o unitÃ¡rio e quantidade inicial em estoque.
*2* - Atualizar Estoque: O usuÃ¡rio pode atualizar a quantidade em estoque de um produto existente.
*3* - Vender Produto: O usuÃ¡rio pode registrar uma venda de um produto, diminuindo a quantidade
      em estoque. Se a quantidade vendida for maior que a quantidade em estoque, 
      informe que nÃ£o Ã© possÃ­vel realizar a venda.
*4* - RelatÃ³rio de Estoque: Exiba um relatÃ³rio com todos os produtos, 
      seus preÃ§os e quantidades em estoque.
'''

estoque = {}  # DicionÃ¡rio para armazenar o estoque (nome do produto: {'preÃ§o': X, 'quantidade': Y})

#definicao de funcoes:
    
def adicionar_produto():
    nome = input("Digite o nome do produto: ")
    price = float(input("Digite o preÃ§o unitÃ¡rio do produto: "))
    quantity = int(input("Digite a quantidade inicial em estoque: "))
    estoque[nome] = {'preÃ§o': price, 'quantidade': quantity}
    print(f"Produto '{nome}' adicionado ao estoque.")

def atualizar_estoque():
    nome = input("Digite o nome do produto para atualizaÃ§Ã£o: ")
    if nome in estoque:
        nova_quantidade = int(input(f"Atualize a quantidade de '{nome}' em estoque: "))
        estoque[nome]['quantidade'] = nova_quantidade
        print(f"Estoque de '{nome}' atualizado para {nova_quantidade}.")
    else:
        print(f"Produto '{nome}' nÃ£o encontrado no estoque.")

def vender_produto():
    nome = input("Digite o nome do produto vendido: ")
    if nome in estoque:
        quantidade_vendida = int(input(f"Digite a quantidade de '{nome}' vendida: "))
        if quantidade_vendida <= estoque[nome]['quantidade']:
            estoque[nome]['quantidade'] -= quantidade_vendida
            print(f"{quantidade_vendida} unidades de '{nome}' vendidas.")
        else:
            print(f"NÃ£o hÃ¡ quantidade suficiente de '{nome}' em estoque para realizar a venda.")
    else:
        print(f"Produto '{nome}' nÃ£o encontrado no estoque.")

def estoque_relatorio():
    print("RelatÃ³rio de Estoque:")
    for product, info in estoque.items():
        print(f"Produto: {product} | PreÃ§o: R${info['preÃ§o']} | Quantidade em Estoque: {info['quantidade']}")


def linha_separacao():
    return print(' -'*30)


# Menu principal
while True:
    linha_separacao()
    print("\n=== Sistema de Gerenciamento de Estoque ===")
    linha_separacao()
    print("1. Adicionar Produto")
    linha_separacao()
    print("2. Atualizar Estoque")
    linha_separacao()
    print("3. Vender Produto")
    linha_separacao()
    print("4. RelatÃ³rio de Estoque")
    linha_separacao()
    print("5. Sair")
    linha_separacao()
    choice = int(input("choice uma opÃ§Ã£o: "))
    
    if choice == 1:
        adicionar_produto()
    elif choice == 2:
        atualizar_estoque()
    elif choice == 3:
        vender_produto()
    elif choice == 4:
        estoque_relatorio()
    elif choice == 5:
        print("Saindo do sistema.")
        break
    else:
        print("OpÃ§Ã£o invÃ¡lida. choice uma opÃ§Ã£o vÃ¡lida.")


'''
ExercÃ­cio 52 - DicionÃ¡rio de FrequÃªncias
Enunciado: Crie uma função que recebe uma lista de palavras e retorna um dicionÃ¡rio 
com a contagem de ocorrÃªncias de cada palavra
'''

def contar_palavras(lista):
    frequencias = {}
    for palavra in lista:
        if palavra in frequencias:
            frequencias[palavra] += 1
        else:
            frequencias[palavra] = 1
    return frequencias



'''
ExercÃ­cio 53 - Escreva uma função que recebe dois dicionÃ¡rios e retorna 
um novo dicionÃ¡rio contendo as chaves e valores de ambos, 
priorizando os valores do segundo dicionÃ¡rio 
em caso de chaves duplicadas
'''

def unir_dicionarios(d1, d2):
    novo_dicionario = d1.copy()
    novo_dicionario.update(d2)
    return novo_dicionario


'''
ExercÃ­cio 54 - função Recursiva em DicionÃ¡rio
Enunciado: Crie uma função que recebe um dicionÃ¡rio aninhado e retorna 
uma lista com todas as chaves presentes no dicionÃ¡rio, incluindo as chaves 
dos dicionÃ¡rios internos.
'''

def listar_chaves(dicionario):
    chaves = []
    for chave, valor in dicionario.items():
        chaves.append(chave)
        if isinstance(valor, dict):
            chaves.extend(listar_chaves(valor))
    return chaves


'''
ExercÃ­cio 55 - função Lambda em DicionÃ¡rio
Enunciado: Crie um dicionÃ¡rio contendo nomes de pessoas como chaves 
e suas idades como valores. Use a função sorted com uma função 
lambda para ordenar o dicionÃ¡rio pelas idades.
'''

pessoas = {
    'Ana': 30,
    'Joao': 25,
    'Maria': 40,
    'Pedro': 22
}

pessoas_ordenadas = dict(sorted(pessoas.items(), key=lambda item: item[1]))


'''
ExercÃ­cio 56 - Aninhamento de funções
Enunciado: Crie uma função que recebe uma lista de nÃºmeros e retorna um dicionÃ¡rio 
com a soma, a mÃ©dia e o desvio padrÃ£o dos nÃºmeros.
'''

import statistics

def analisar_numeros(lista):
    resultado = {
        'soma': sum(lista),
        'media': statistics.mean(lista),
        'desvio_padrao': statistics.stdev(lista)
    }
    return resultado


'''
ExercÃ­cio 57 - função que Aceita NÃºmero VariÃ¡vel de Argumentos
Enunciado: Crie uma função que recebe um nÃºmero variÃ¡vel de argumentos de palavras
e retorna um dicionÃ¡rio com a contagem de ocorrÃªncias de cada palavra.
'''

def contar_palavras_variaveis(*args):
    frequencias = {}
    for palavra in args:
        if palavra in frequencias:
            frequencias[palavra] += 1
        else:
            frequencias[palavra] = 1
    return frequencias


'''
ExercÃ­cio 58 - função que Retorna função
Enunciado: Crie uma função que recebe um valor inteiro n e 
retorna uma função que recebe um nÃºmero e retorna o seu 
dobro elevado a n.
'''

def criar_funcao_potencia(n):
    def funcao_potencia(x):
        return (2 * x) ** n
    return funcao_potencia


'''
ExercÃ­cio 59 - Contagem de Letras em Texto
Enunciado: Escreva uma função que recebe um texto e retorna um 
dicionÃ¡rio com a contagem de cada letra presente no texto.
'''

def contar_letras(texto):
    frequencias = {}
    for letra in texto:
        if letra.isalpha():
            letra = letra.lower()
            if letra in frequencias:
                frequencias[letra] += 1
            else:
                frequencias[letra] = 1
    return frequencias


'''
ExercÃ­cio 60 - função Decoradora com DicionÃ¡rio
Enunciado: Crie uma função decoradora que recebe um dicionÃ¡rio de argumentos e seus 
valores padrÃ£o. 
A decoradora deve ser usada para adicionar argumentos opcionais a uma função.
'''

def decorador_argumentos_padrao(argumentos_padrao):
    def decorator(func):
        def wrapper(*args, **kwargs):
            for arg, valor in argumentos_padrao.items():
                if arg not in kwargs:
                    kwargs[arg] = valor
            return func(*args, **kwargs)
        return wrapper
    return decorator

# Uso da decoradora
argumentos_padrao = {'x': 0, 'y': 0}
@decorador_argumentos_padrao(argumentos_padrao)
def minha_funcao(a, b, x, y):
    return a + b + x + y


'''
ExercÃ­cio 61 - AnÃ¡lise de Texto com função e DicionÃ¡rio
Enunciado: Crie uma função que recebe um texto e retorna um dicionÃ¡rio contendo a 
contagem de palavras Ãºnicas e a contagem de cada palavra repetida no texto.
'''

def analisar_texto(texto):
    palavras = texto.split()
    contagem_palavras = {}
    for palavra in palavras:
        palavra = palavra.lower()
        if palavra in contagem_palavras:
            contagem_palavras[palavra] += 1
        else:
            contagem_palavras[palavra] = 1
    return {
        'palavras_unicas': len(contagem_palavras),
        'contagem_palavras': contagem_palavras
    }


'''
ExercÃ­cio 62 - Sistema de GestÃ£o de Projetos

Enunciado:
Imagine que vocÃª estÃ¡ desenvolvendo um sistema de gestÃ£o de projetos para uma equipe 
de engenheiros de software. VocÃª foi solicitado a criar um mÃ³dulo em Python que 
permita aos engenheiros registrarem suas atividades diÃ¡rias e depois analisarem os 
dados para obter insights sobre o progresso do projeto. VocÃª deve usar funções 
e dicionÃ¡rios para implementar essa funcionalidade.

Crie uma função registrar_atividade que permita aos engenheiros registrarem suas atividades diÃ¡rias. 
A função deve receber como argumentos o nome do engenheiro, a data da atividade e uma 
descriÃ§Ã£o da atividade. As atividades devem ser armazenadas em um dicionÃ¡rio onde as chaves sÃ£o 
as datas e os valores sÃ£o listas de atividades.

Crie uma função analise_progresso que permita aos engenheiros analisarem o progresso do projeto. 
A função deve receber como argumento o dicionÃ¡rio de atividades e calcular as seguintes mÃ©tricas:

Total de atividades realizadas por cada engenheiro.
Total de atividades realizadas por data.
MÃ©dia de atividades realizadas por engenheiro.
MÃ©dia de atividades realizadas por dia.
Implemente um menu interativo que permita aos usuÃ¡rios escolherem entre registrar uma nova 
atividade ou realizar uma anÃ¡lise de progresso. O menu deve ser implementado em um loop contÃ­nuo 
atÃ© que o usuÃ¡rio escolha sair do programa.
'''

def registrar_atividade(atividades, engenheiro, data, descricao):
    if data in atividades:
        atividades[data].append((engenheiro, descricao))
    else:
        atividades[data] = [(engenheiro, descricao)]

def analise_progresso(atividades):
    total_por_engenheiro = {}
    total_por_data = {}
    for data, lista_atividades in atividades.items():
        total_por_data[data] = len(lista_atividades)
        for engenheiro, _ in lista_atividades:
            if engenheiro in total_por_engenheiro:
                total_por_engenheiro[engenheiro] += 1
            else:
                total_por_engenheiro[engenheiro] = 1
    
    media_por_engenheiro = sum(total_por_engenheiro.values()) / len(total_por_engenheiro)
    media_por_data = sum(total_por_data.values()) / len(total_por_data)
    
    return {
        'total_por_engenheiro': total_por_engenheiro,
        'total_por_data': total_por_data,
        'media_por_engenheiro': media_por_engenheiro,
        'media_por_data': media_por_data
    }

atividades = {}
while True:
    print("Menu:")
    print("1. Registrar atividade")
    print("2. AnÃ¡lise de progresso")
    print("3. Sair")
    
    escolha = int(input("Escolha uma opÃ§Ã£o: "))
    
    if escolha == 1:
        engenheiro = input("Nome do engenheiro: ")
        data = input("Data da atividade: ")
        descricao = input("DescriÃ§Ã£o da atividade: ")
        registrar_atividade(atividades, engenheiro, data, descricao)
        print("Atividade registrada com sucesso.")
    elif escolha == 2:
        resultado_analise = analise_progresso(atividades)
        print("AnÃ¡lise de Progresso:")
        print("Total de atividades por engenheiro:", resultado_analise['total_por_engenheiro'])
        print("Total de atividades por data:", resultado_analise['total_por_data'])
        print("MÃ©dia de atividades por engenheiro:", resultado_analise['media_por_engenheiro'])
        print("MÃ©dia de atividades por data:", resultado_analise['media_por_data'])
    elif escolha == 3:
        print("Saindo do programa.")
        break
    else:
        print("OpÃ§Ã£o invÃ¡lida. Escolha novamente.")


'''
Exercicios com Sets - Conjuntos

ExercÃ­cio 63 - UniÃ£o de Conjuntos:
Enunciado: Crie dois conjuntos A e B e retorne a uniÃ£o deles.
'''

A = {1, 2, 3}
B = {3, 4, 5}
uniao = A.union(B)
print(uniao)

    
'''
ExercÃ­cio 64 - InterseÃ§Ã£o de Conjuntos:
Enunciado: Crie dois conjuntos A e B e retorne a interseÃ§Ã£o deles.
'''

A = {1, 2, 3}
B = {3, 4, 5}
diferenca = A.difference(B)
print(diferenca)
    

'''
ExercÃ­cio 65 - VerificaÃ§Ã£o de Subconjunto:
Enunciado: Crie dois conjuntos A e B e verifique se A Ã© um subconjunto de B.
'''

A = {1, 2}
B = {1, 2, 3, 4, 5}
eh_subconjunto = A.issubset(B)
print(eh_subconjunto)


'''
ExercÃ­cio 66 - Adicionar e Remover Elementos:
Enunciado: Crie um conjunto vazio, adicione alguns elementos e depois remova um deles.
'''

conjunto = set()
conjunto.add(1)
conjunto.add(2)
conjunto.add(3)
conjunto.remove(2)
print(conjunto)


'''
ExercÃ­cio 67 - ConcatenaÃ§Ã£o de Conjuntos:
Enunciado: Crie dois conjuntos A e B e concatene-os em um terceiro conjunto C.
'''

A = {1, 2, 3}
B = {3, 4, 5}
C = A | B
print(C)


'''
ExercÃ­cio 68 - RemoÃ§Ã£o de Duplicatas em uma Lista:
Enunciado: Dada uma lista com elementos duplicados, crie um conjunto com os elementos Ãºnicos.
'''

lista = [1, 2, 2, 3, 4, 4, 5]
conjunto = set(lista)
print(conjunto)



'''
ExercÃ­cio 69 - VerificaÃ§Ã£o de Igualdade de Conjuntos:
Enunciado: Crie dois conjuntos A e B e verifique se eles sÃ£o iguais.
'''

A = {1, 2, 3}
B = {3, 2, 1}
sao_iguais = A == B
print(sao_iguais)


'''
ExercÃ­cio 70 - Conjunto de PotÃªncias:
Enunciado: Crie um conjunto de potÃªncias de um conjunto A dado.
'''


A = {1, 2, 3}
potencias = {frozenset({x for x in A if (i & (1 << A.index(x))) != 0}) for i in range(1 << len(A))}
print(potencias)


'''
ExercÃ­cio 71 - RemoÃ§Ã£o de Elementos Repetidos em Conjunto:
Enunciado: Dado um conjunto com elementos repetidos, remova as duplicatas.
'''

conjunto = {1, 2, 2, 3, 3, 4, 4, 5}
conjunto_sem_duplicatas = set()
for elemento in conjunto:
    conjunto_sem_duplicatas.add(elemento)
print(conjunto_sem_duplicatas)


'''
ExercÃ­cio 72 - Desafio - AnÃ¡lise de Dados de Compras:
Enunciado: VocÃª recebeu um conjunto de dados contendo informaÃ§Ãµes sobre compras em uma loja. 
Cada compra Ã© representada como uma tupla contendo o nome do cliente e os itens comprados. 
Seu objetivo Ã© realizar algumas anÃ¡lises com base nesses dados.

VocÃª deve implementar as seguintes funções:

total_gasto_por_cliente(dados): Recebe um conjunto de dados de compras e retorna um dicionÃ¡rio
 onde as chaves sÃ£o os nomes dos clientes e os valores sÃ£o o total gasto por cada cliente.

clientes_que_compraram(item, dados): Recebe um item especÃ­fico e um conjunto de dados 
de compras e retorna um conjunto com os nomes dos clientes que compraram esse item.

itens_universais(dados): Recebe um conjunto de dados de compras e retorna um conjunto 
com os itens que foram comprados por todos os clientes.
'''

dados_de_compras = [
    ("Joao", {"Camiseta", "Calça", "Tenis"}),
    ("Maria", {"Calça", "óculos de Sol", "Relógio"}),
    ("João", {"Camiseta", "Tenis"}),
    ("Ana", {"Camiseta", "Calça", "óculos de Sol"}),
    ("Pedro", {"Tenis", "Relógio"}),
]

# Exemplo de uso das funções
dados_de_compras = [
    ("Joao", {"Camiseta", "Calça", "Tenis"}),
    ("Maria", {"Calça", "Óculos de Sol", "Relógio"}),
    ("João", {"Camiseta", "Tenis"}),
    ("Ana", {"Camiseta", "Calça", "óculos de Sol"}),
    ("Pedro", {"Tenis", "Relógio"}),
]

# Teste da primeira função
total_gasto = total_gasto_por_cliente(dados_de_compras)
print(total_gasto)  # Deve imprimir {'Joao': 3, 'Maria': 3, 'João': 2, 'Ana': 3, 'Pedro': 2}

# Teste da segunda função
compradores = clientes_que_compraram("Calça", dados_de_compras)
print(compradores)  # Deve imprimir {'Joao', 'Maria', 'Ana'}

# Teste da terceira função
itens_universais = itens_universais(dados_de_compras)
print(itens_universais)  # Deve imprimir {'Tenis'}

'''
Exercicio 73 - Sistema de Gerenciamento de Tarefas

Enunciado: Crie um sistema de gerenciamento de tarefas em Python. Implemente classes para 
representar tarefas, uma lista de tarefas e funções para adicionar tarefas, 
marcar tarefas como concluí­das e listar tarefas pendentes.
'''

class Tarefa:
    def __init__(self, descricao, concluida=False):
        self.descricao = descricao
        self.concluida = concluida

class ListaDeTarefas:
    def __init__(self):
        self.tarefas = []

    def adicionar_tarefa(self, tarefa):
        self.tarefas.append(tarefa)

    def marcar_tarefa_como_concluida(self, indice):
        if 0 <= indice < len(self.tarefas):
            self.tarefas[indice].concluida = True

    def listar_tarefas_pendentes(self):
        tarefas_pendentes = [tarefa for tarefa in self.tarefas if not tarefa.concluida]
        return [tarefa.descricao for tarefa in tarefas_pendentes]

# Exemplo de uso:
lista = ListaDeTarefas()
lista.adicionar_tarefa(Tarefa("Lavar o carro"))
lista.adicionar_tarefa(Tarefa("Fazer compras"))
lista.marcar_tarefa_como_concluida(1)
print("Tarefas pendentes:", lista.listar_tarefas_pendentes())

'''
Exercicio 74 -  Sistema de Gerenciamento de FinanÃ§as Pessoais
Enunciado: Desenvolva um sistema de gerenciamento de finanÃ§as pessoais em Python. 
Crie classes para representar transaÃ§Ãµes financeiras e contas bancÃ¡rias, 
e implemente funções para adicionar transaÃ§Ãµes, calcular o saldo e listar transaÃ§Ãµes de uma conta.
'''

class Transacao:
    def __init__(self, descricao, valor):
        self.descricao = descricao
        self.valor = valor

class ContaBancaria:
    def __init__(self):
        self.saldo = 0
        self.transacoes = []

    def adicionar_transacao(self, transacao):
        self.transacoes.append(transacao)
        self.saldo += transacao.valor

    def listar_transacoes(self):
        return [(t.descricao, t.valor) for t in self.transacoes]

# Exemplo de uso:
conta = ContaBancaria()
conta.adicionar_transacao(Transacao("SalÃ¡rio", 3000.0))
conta.adicionar_transacao(Transacao("Compras", -500.0))
print("Saldo:", conta.saldo)
print("TransaÃ§Ãµes:", conta.listar_transacoes())


'''
Exercicio 75 - Sistema de Biblioteca
Enunciado: Crie um sistema de biblioteca em Python. Implemente classes para representar livros, 
usuÃ¡rios e emprÃ©stimos. Desenvolva funções para emprestar e devolver livros, 
listar livros disponÃ­veis e gerar relatÃ³rios de emprÃ©stimos.
'''

class Livro:
    def __init__(self, titulo, autor):
        self.titulo = titulo
        self.autor = autor
        self.disponivel = True

class Usuario:
    def __init__(self, nome):
        self.nome = nome

class Biblioteca:
    def __init__(self):
        self.livros = []
        self.usuarios = []
        self.emprestimos = []

    def emprestar_livro(self, livro, usuario):
        if livro.disponivel:
            livro.disponivel = False
            self.emprestimos.append((livro, usuario))

    def devolver_livro(self, livro, usuario):
        if (livro, usuario) in self.emprestimos:
            livro.disponivel = True
            self.emprestimos.remove((livro, usuario))

    def listar_livros_disponiveis(self):
        return [livro.titulo for livro in self.livros if livro.disponivel]

# Exemplo de uso:
biblioteca = Biblioteca()
livro1 = Livro("Aprendendo Python", "John Smith")
usuario1 = Usuario("Alice")
biblioteca.livros.append(livro1)
biblioteca.usuarios.append(usuario1)
biblioteca.emprestar_livro(livro1, usuario1)
print("Livros disponÃ­veis:", biblioteca.listar_livros_disponiveis())

'''
Exercicio 76 - : Sistema de Reservas de Hotel
Enunciado: Crie um sistema de reservas de hotel em Python. 
Implemente classes para representar quartos, reservas e hÃ³spedes. 
Desenvolva funções para fazer reservas, verificar disponibilidade, 
listar reservas e calcular a receita total do hotel.
'''

class Quarto:
    def __init__(self, numero, capacidade, preco_diaria):
        self.numero = numero
        self.capacidade = capacidade
        self.preco_diaria = preco_diaria
        self.disponivel = True

class Reserva:
    def __init__(self, quarto, hospede, data_inicio, data_fim):
        self.quarto = quarto
        self.hospede = hospede
        self.data_inicio = data_inicio
        self.data_fim = data_fim

class Hospede:
    def __init__(self, nome):
        self.nome = nome

class Hotel:
    def __init__(self):
        self.quartos = []
        self.reservas = []

    def fazer_reserva(self, quarto, hospede, data_inicio, data_fim):
        if quarto.disponivel:
            quarto.disponivel = False
            self.reservas.append(Reserva(quarto, hospede, data_inicio, data_fim))

    def verificar_disponibilidade(self, data_inicio, data_fim):
        quartos_disponiveis = [quarto for quarto in self.quartos if quarto.disponivel]
        return [quarto.numero for quarto in quartos_disponiveis]

    def calcular_receita_total(self):
        return sum(reserva.quarto.preco_diaria for reserva in self.reservas)

# Exemplo de uso:
hotel = Hotel()
quarto1 = Quarto(101, 2, 100.0)
hospede1 = Hospede("Joao")
hotel.quartos.append(quarto1)
hotel.fazer_reserva(quarto1, hospede1, "2023-09-15", "2023-09-18")
print("Quartos disponÃ­veis:", hotel.verificar_disponibilidade("2023-09-15", "2023-09-18"))
print("Receita Total:", hotel.calcular_receita_total())

'''
Exercicio 77 - : Sistema de Gerenciamento de FuncionÃ¡rios
Enunciado: Desenvolva um sistema de gerenciamento de funcionÃ¡rios em Python. Crie classes para representar
funcionÃ¡rios, departamentos e cargos. Implemente funções para adicionar funcionÃ¡rios, 
calcular salÃ¡rios, atribuir cargos e gerar relatÃ³rios de funcionÃ¡rios.
'''

class Funcionario:
    def __init__(self, nome, salario_base):
        self.nome = nome
        self.salario_base = salario_base

class Cargo:
    def __init__(self, nome, salario_adicional):
        self.nome = nome
        self.salario_adicional = salario_adicional

class Departamento:
    def __init__(self, nome):
        self.nome = nome
        self.funcionarios = []

    def adicionar_funcionario(self, funcionario):
        self.funcionarios.append(funcionario)

    def calcular_salario_total(self):
        return sum(funcionario.salario_base for funcionario in self.funcionarios)

# Exemplo de uso:
departamento = Departamento("Vendas")
cargo1 = Cargo("Vendedor", 500.0)
funcionario1 = Funcionario("Maria", 2000.0)
departamento.adicionar_funcionario(funcionario1)
print("SalÃ¡rio Total do Departamento:", departamento.calcular_salario_total())

'''
Exercicio 78 - Sistema de Gerenciamento de InventÃ¡rio de TI
Enunciado: Crie um sistema de gerenciamento de inventÃ¡rio de TI em Python. 
Implemente classes para representar dispositivos, manutenÃ§Ãµes e custos. 
Desenvolva funções para adicionar dispositivos, registrar manutenÃ§Ãµes, 
calcular custos de manutenÃ§Ã£o e gerar relatÃ³rios de inventÃ¡rio.
'''
class DispositivoTI:
    def __init__(self, nome, custo):
        self.nome = nome
        self.custo = custo

class Manutencao:
    def __init__(self, dispositivo, descricao, custo):
        self.dispositivo = dispositivo
        self.descricao = descricao
        self.custo = custo

class InventarioTI:
    def __init__(self):
        self.dispositivos = []
        self.manutencoes = []

    def adicionar_dispositivo(self, dispositivo):
        self.dispositivos.append(dispositivo)

    def registrar_manutencao(self, dispositivo, descricao, custo):
        self.manutencoes.append(Manutencao(dispositivo, descricao, custo))

    def calcular_custo_total_manutencoes(self):
        return sum(manutencao.custo for manutencao in self.manutencoes)

# Exemplo de uso:
inventario = InventarioTI()
dispositivo1 = DispositivoTI("Servidor", 5000.0)
inventario.adicionar_dispositivo(dispositivo1)
inventario.registrar_manutencao(dispositivo1, "SubstituiÃ§Ã£o de disco", 200.0)
print("Custo Total de ManutenÃ§Ãµes:", inventario.calcular_custo_total_manutencoes())

'''
Exercicio 79 -  Sistema de Reservas de Voos
Enunciado: Desenvolva um sistema de reservas de voos em Python. 
Implemente classes para representar voos, passageiros e reservas. 
Desenvolva funções para fazer reservas, verificar disponibilidade, listar reservas e calcular
a receita total da companhia aÃ©rea.
'''

class Voo:
    def __init__(self, numero, capacidade, preco_passagem):
        self.numero = numero
        self.capacidade = capacidade
        self.preco_passagem = preco_passagem
        self.disponivel = True

class ReservaVoo:
    def __init__(self, voo, passageiro):
        self.voo = voo
        self.passageiro = passageiro

class Passageiro:
    def __init__(self, nome):
        self.nome = nome

class CompanhiaAerea:
    def __init__(self):
        self.voos = []
        self.reservas = []

    def fazer_reserva_voo(self, voo, passageiro):
        if voo.disponivel:
            voo.disponivel = False
            self.reservas.append(ReservaVoo(voo, passageiro))

    def verificar_disponibilidade_voo(self):
        voos_disponiveis = [voo for voo in self.voos if voo.disponivel]
        return [voo.numero for voo in voos_disponiveis]

    def calcular_receita_total(self):
        return sum(reserva.voo.preco_passagem for reserva in self.reservas)

# Exemplo de uso:
companhia = CompanhiaAerea()
voo1 = Voo("AA101", 150, 300.0)
passageiro1 = Passageiro("João")
companhia.voos.append(voo1)
companhia.fazer_reserva_voo(voo1, passageiro1)
print("Voos disponÃ­veis:", companhia.verificar_disponibilidade_voo())
print("Receita Total da Companhia AÃ©rea:", companhia.calcular_receita_total())

'''
Exercicio 80 -  Sistema de Gerenciamento de Projetos
Enunciado: Crie um sistema de gerenciamento de projetos em Python. 
Implemente classes para representar projetos, tarefas e equipes de trabalho. 
Desenvolva funções para criar projetos, atribuir tarefas a projetos, acompanhar o progresso
das tarefas e gerar relatÃ³rios de status de projeto.
'''

class Projeto:
    def __init__(self, nome, equipe):
        self.nome = nome
        self.equipe = equipe
        self.tarefas = []

    def adicionar_tarefa(self, tarefa):
        self.tarefas.append(tarefa)

    def calcular_progresso(self):
        if not self.tarefas:
            return 0.0
        tarefas_concluidas = sum(1 for tarefa in self.tarefas if tarefa.concluida)
        return (tarefas_concluidas / len(self.tarefas)) * 100.0

class Tarefa:
    def __init__(self, descricao, concluida=False):
        self.descricao = descricao
        self.concluida = concluida

class Equipe:
    def __init__(self, nome):
        self.nome = nome

# Exemplo de uso:
equipe1 = Equipe("Desenvolvimento")
projeto1 = Projeto("Website", equipe1)
projeto1.adicionar_tarefa(Tarefa("Design da PÃ¡gina Inicial"))
projeto1.adicionar_tarefa(Tarefa("Implementar Funcionalidades"))
projeto1.tarefas[1].concluida = True
print("Progresso do Projeto:", projeto1.calcular_progresso(), "%")

'''
Exercicio 81 - Sistema de Gerenciamento de Pedidos de Vendas
Enunciado: Desenvolva um sistema de gerenciamento de pedidos de vendas em Python. 
Implemente classes para representar pedidos, produtos e clientes. 
Desenvolva funções para adicionar pedidos, atualizar status de pedidos, calcular o total de vendas
por cliente e gerar relatÃ³rios de vendas.
'''

class Pedido:
    def __init__(self, numero, cliente, produtos):
        self.numero = numero
        self.cliente = cliente
        self.produtos = produtos
        self.status = "Pendente"

    def atualizar_status(self, novo_status):
        self.status = novo_status

class Produto:
    def __init__(self, nome, preco):
        self.nome = nome
        self.preco = preco

class Cliente:
    def __init__(self, nome):
        self.nome = nome

class SistemaDeVendas:
    def __init__(self):
        self.pedidos = []

    def adicionar_pedido(self, pedido):
        self.pedidos.append(pedido)

    def calcular_total_vendas_por_cliente(self):
        vendas_por_cliente = {}
        for pedido in self.pedidos:
            cliente = pedido.cliente.nome
            if cliente in vendas_por_cliente:
                vendas_por_cliente[cliente] += sum(produto.preco for produto in pedido.produtos)
            else:
                vendas_por_cliente[cliente] = sum(produto.preco for produto in pedido.produtos)
        return vendas_por_cliente

# Exemplo de uso:
sistema_vendas = SistemaDeVendas()
cliente1 = Cliente("Maria")
produto1 = Produto("Laptop", 1000.0)
produto2 = Produto("Impressora", 200.0)
pedido1 = Pedido(1, cliente1, [produto1, produto2])
sistema_vendas.adicionar_pedido(pedido1)
print("Total de Vendas por Cliente:", sistema_vendas.calcular_total_vendas_por_cliente())

'''
Exercicio 82 - : Sistema de Controle de Estoque AvanÃ§ado

Enunciado: Crie um sistema avanÃ§ado de controle de estoque em Python. 
Implemente classes para representar produtos, categorias, fornecedores e estoque. 
Desenvolva funções para adicionar produtos ao estoque, registrar fornecedores, 
calcular o valor total do estoque e listar produtos por categoria.
'''

class Produto:
    def __init__(self, nome, categoria, preco, quantidade):
        self.nome = nome
        self.categoria = categoria
        self.preco = preco
        self.quantidade = quantidade

class Categoria:
    def __init__(self, nome):
        self.nome = nome

class Fornecedor:
    def __init__(self, nome):
        self.nome = nome
        self.produtos_fornecidos = []

    def registrar_produto(self, produto):
        self.produtos_fornecidos.append(produto)

class Estoque:
    def __init__(self):
        self.produtos = []
        self.fornecedores = []

    def adicionar_produto(self, produto):
        self.produtos.append(produto)

    def registrar_fornecedor(self, fornecedor):
        self.fornecedores.append(fornecedor)

    def calcular_valor_total_estoque(self):
        return sum(produto.preco * produto.quantidade for produto in self.produtos)

    def listar_produtos_por_categoria(self, categoria):
        return [produto.nome for produto in self.produtos if produto.categoria.nome == categoria]

# Exemplo de uso:
estoque = Estoque()
categoria1 = Categoria("EletrÃ´nicos")
fornecedor1 = Fornecedor("TechCo")
produto1 = Produto("Laptop", categoria1, 1000.0, 10)
fornecedor1.registrar_produto(produto1)
estoque.adicionar_produto(produto1)
estoque.registrar_fornecedor(fornecedor1)
print("Valor Total do Estoque:", estoque.calcular_valor_total_estoque())
print("Produtos EletrÃ´nicos:", estoque.listar_produtos_por_categoria("EletrÃ´nicos"))


'''
Exercicios com Matrizes
'''

'''
Exercicio 83 - Escreva uma função chamada soma_matrizes que recebe duas matrizes de
nÃºmeros inteiros como parÃ¢metros e retorna a matriz resultante da soma das duas matrizes. 
Se as matrizes nÃ£o tiverem o mesmo tamanho, a função deve retornar None. 
Por exemplo, se A = [[1, 2], [3, 4]] e B = [[5, 6], [7, 8]], entÃ£o soma_matrizes(A, B) 
deve retornar [[6, 8], [10, 12]].'
'Python
Esse cÃ³digo Ã© gerado por IA. Examine e use com cuidado. Visite nossas Perguntas Frequentes para obter mais informaÃ§Ãµes.
'''

def soma_matrizes(A, B):
  # Verifica se as matrizes tÃªm o mesmo tamanho
  if len(A) != len(B) or len(A[0]) != len(B[0]):
    return None
  # Cria uma matriz vazia para armazenar o resultado
  C = []
  # Percorre as linhas das matrizes
  for i in range(len(A)):
    # Cria uma lista vazia para armazenar a linha resultante
    linha = []
    # Percorre as colunas das matrizes
    for j in range(len(A[0])):
      # Soma os elementos correspondentes das matrizes e adiciona Ã  lista
      linha.append(A[i][j] + B[i][j])
    # Adiciona a lista Ã  matriz resultante
    C.append(linha)
  # Retorna a matriz resultante
  return C

'''
Exercico 84 Escreva uma função chamada transposta que recebe uma matriz de nÃºmeros inteiros 
como parÃ¢metro e retorna a sua transposta, ou seja, a matriz obtida trocando as linhas 
pelas colunas. Por exemplo, se A = [[1, 2], [3, 4]], entÃ£o transposta(A) deve 
retornar [[1, 3], [2, 4]].
'''

def transposta(A):
  # Cria uma matriz vazia para armazenar o resultado
  B = []
  # Percorre as colunas da matriz original
  for j in range(len(A[0])):
    # Cria uma lista vazia para armazenar a linha resultante
    linha = []
    # Percorre as linhas da matriz original
    for i in range(len(A)):
      # Adiciona o elemento correspondente da matriz original Ã  lista
      linha.append(A[i][j])
    # Adiciona a lista Ã  matriz resultante
    B.append(linha)
  # Retorna a matriz resultante
  return B

'''
Exercicio 85 - Escreva uma função chamada multiplica_matrizes que recebe duas matrizes de 
nÃºmeros inteiros como parÃ¢metros e retorna a matriz resultante da multiplicaÃ§Ã£o das duas matrizes. 
Se as matrizes nÃ£o forem compatÃ­veis para a multiplicaÃ§Ã£o, a função deve retornar None. 
Por exemplo, se A = [[1, 2], [3, 4]] e B = [[5, 6], [7, 8]], entÃ£o multiplica_matrizes(A, B) 
deve retornar [[19, 22], [43, 50]].
'''

def multiplica_matrizes(A, B):
  # Verifica se as matrizes sÃ£o compatÃ­veis para a multiplicaÃ§Ã£o
  if len(A[0]) != len(B):
    return None
  # Cria uma matriz vazia para armazenar o resultado
  C = []
  # Percorre as linhas da primeira matriz
  for i in range(len(A)):
    # Cria uma lista vazia para armazenar a linha resultante
    linha = []
    # Percorre as colunas da segunda matriz
    for j in range(len(B[0])):
      # Inicializa uma variÃ¡vel para armazenar o elemento resultante
      elemento = 0
      # Percorre as colunas da primeira matriz ou as linhas da segunda matriz (sÃ£o iguais)
      for k in range(len(A[0])):
        # Multiplica os elementos correspondentes das matrizes e adiciona Ã  variÃ¡vel
        elemento += A[i][k] * B[k][j]
      # Adiciona o elemento Ã  lista
      linha.append(elemento)
    # Adiciona a lista Ã  matriz resultante
    C.append(linha)
  # Retorna a matriz resultante
  return C

'''
Exercicio 86 - Escreva uma função chamada determinante que recebe uma matriz quadrada de nÃºmeros 
inteiros como parÃ¢metro e retorna o seu determinante, ou seja, um nÃºmero que representa a Ã¡rea do 
paralelogramo formado pelos vetores da matriz. Se a matriz nÃ£o for quadrada, a função deve retornar None. Por exemplo, se A = [[1, 2], [3, 4]], entÃ£o determinante(A) deve retornar -2.
'''
def determinante(A):
  # Verifica se a matriz Ã© quadrada
  if len(A) != len(A[0]):
    return None
  # Se a matriz Ã© de ordem 1, o determinante Ã© o prÃ³prio elemento
  if len(A) == 1:
    return A[0][0]
  # Se a matriz Ã© de ordem 2, o determinante Ã© dado pela regra de Sarrus
  if len(A) == 2:
    return A[0][0] * A[1][1] - A[0][1] * A[1][0]
  # Se a matriz Ã© de ordem maior que 2, o determinante Ã© dado pela expansÃ£o de Laplace
  # Escolhe a primeira linha para fazer a expansÃ£o
  det = 0 # Inicializa uma variÃ¡vel para armazenar o determinante
  for j in range(len(A[0])): # Percorre as colunas da primeira linha
    # Calcula o cofator do elemento A[0][j]
    cofator = (-1)**(j) * A[0][j] * determinante(submatriz(A, 0, j))
    # Adiciona o cofator ao determinante
    det += cofator
  # Retorna o determinante
  return det

# função auxiliar para calcular a submatriz resultante da remoÃ§Ã£o de uma linha e uma coluna de uma matriz
def submatriz(A, i, j):
  # Cria uma matriz vazia para armazenar o resultado
  B = []
  # Percorre as linhas da matriz original
  for k in range(len(A)):
    # Ignora a linha i
    if k != i:
      # Cria uma lista vazia para armazenar a linha resultante
      linha = []
      # Percorre as colunas da matriz original
      for l in range(len(A[0])):
        # Ignora a coluna j
        if l != j:
          # Adiciona o elemento correspondente da matriz original Ã  lista
          linha.append(A[k][l])
      # Adiciona a lista Ã  matriz resultante
      B.append(linha)
  # Retorna a matriz resultante
  return B


'''
Exercicio 87 - Enunciado: Crie uma função em Python que multiplica uma matriz 
por um escalar.
'''

def multiplica_matriz_por_escalar(matriz, escalar):
    resultado = []
    for i in range(len(matriz)):
        linha = []
        for j in range(len(matriz[0])):
            linha.append(matriz[i][j] * escalar)
        resultado.append(linha)
    return resultado

'''
Exercicio 88 -  Implemente uma função que retorna a matriz transposta de uma matriz dada.
nota: Define-se como matriz transposta uma matriz qualquer resultante da troca ordenada
das linhas pelas colunas de uma matriz chamada de original. 
Matematicamente, uma transposta de uma matriz Ã© representada por: A = At. 
Antes, porÃ©m, Ã© preciso compreender o conceito de matrizes.
'''
def matriz_transposta(matriz):
    transposta = []
    for j in range(len(matriz[0])):
        linha = []
        for i in range(len(matriz)):
            linha.append(matriz[i][j])
        transposta.append(linha)
    return transposta

'''
Exercicio 89 - Escreva uma função que calcula o produto de duas matrizes.

'''

def produto_de_matrizes(matriz1, matriz2):
    if len(matriz1[0]) != len(matriz2):
        return None  # NÃ£o Ã© possÃ­vel multiplicar as matrizes
    resultado = []
    for i in range(len(matriz1)):
        linha = []
        for j in range(len(matriz2[0])):
            soma = 0
            for k in range(len(matriz2)):
                soma += matriz1[i][k] * matriz2[k][j]
            linha.append(soma)
        resultado.append(linha)
    return resultado

'''
Exercicio 90 - Crie uma função que verifica se uma matriz Ã© simÃ©trica 
(ou seja, igual Ã  sua transposta).

'''

def matriz_simetrica(matriz):
    transposta = matriz_transposta(matriz)
    return matriz == transposta

'''
Exercicio 91 -  Implemente uma função que retorne a diagonal principal de uma matriz.

'''

def diagonal_principal(matriz):
    diagonal = []
    for i in range(len(matriz)):
        diagonal.append(matriz[i][i])
    return diagonal

'''
Exercicio 92 - Escreva um programa que verifica se uma matriz Ã© uma matriz identidade.

'''

def matriz_identidade(matriz):
    for i in range(len(matriz)):
        for j in range(len(matriz[0])):
            if i == j:
                if matriz[i][j] != 1:
                    return False
            else:
                if matriz[i][j] != 0:
                    return False
    return True

'''
Exercicio 93 - Crie uma função que calcula a soma de todos os elementos de uma matriz.

'''

def soma_elementos_matriz(matriz):
    soma = 0
    for linha in matriz:
        soma += sum(linha)
    return soma

'''
Exercicio 94 -  Implemente uma função que verifica se uma matriz Ã© esparsa 
(maioria dos elementos Ã© zero).

'''

def matriz_esparsa(matriz):
    count_zeros = 0
    total_elements = len(matriz) * len(matriz[0])
    for linha in matriz:
        count_zeros += linha.count(0)
    return count_zeros > (total_elements / 2)

'''
Exercicio 95 - Escreva um programa em Python que leia uma matriz quadrada de ordem n e 
imprima a sua diagonal principal e a sua diagonal secundÃ¡ria. 
A entrada deve ser o valor de n seguido dos elementos da matriz, um por linha. 
A saÃ­da deve ser as duas diagonais, uma por linha, separadas por um espaÃ§o em branco. 
Por exemplo:
Entrada:
3
1
2
3
4
5
6
7
8
9

Saida
1 5 9
3 5 7
'''


# LÃª o valor de n
n = int(input())

# Cria uma matriz vazia
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(n):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria duas listas vazias para armazenar as diagonais
diagonal_principal = []
diagonal_secundaria = []

# Percorre a matriz e extrai as diagonais
for i in range(n):
  for j in range(n):
    # Se i == j, o elemento pertence Ã  diagonal principal
    if i == j:
      diagonal_principal.append(matriz[i][j])
    # Se i + j == n - 1, o elemento pertence Ã  diagonal secundÃ¡ria
    if i + j == n - 1:
      diagonal_secundaria.append(matriz[i][j])

# Imprime as diagonais, separadas por um espaÃ§o em branco
print(*diagonal_principal)
print(*diagonal_secundaria)
    

'''
Exercicio 96 - Escreva um programa em Python que leia duas matrizes de nÃºmeros inteiros, 
A e B, de dimensÃµes m x n e n x p, respectivamente, e calcule e imprima o produto 
matricial A x B. 
A entrada deve ser os valores de m, n e p, seguidos dos elementos das matrizes, 
um por linha. A saÃ­da deve ser a matriz resultante, com os elementos separados por
 um espaÃ§o em branco. Por exemplo:
2
3
4
1
2
3
4
5
6
1
0
0
0
0
1
0
0
0
0
1
0
0
0
0
1

Saida
1 2 3 4
4 5 6 6

'''

# LÃª os valores de m, n e p
m = int(input())
n = int(input())
p = int(input())

# Cria duas matrizes vazias para armazenar A e B
A = []
B = []

# LÃª os elementos da matriz A
for i in range(m):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(n):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz A
  A.append(linha)

# LÃª os elementos da matriz B
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(p):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz B
  B.append(linha)

# Cria uma matriz vazia para armazenar o produto A x B
C = []

# Calcula o produto A x B
for i in range(m):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(p):
    # Inicializa a soma com zero
    soma = 0
    # Multiplica a i-Ã©sima linha de A pela j-Ã©sima coluna de B e acumula na soma
    for k in range(n):
      soma += A[i][k] * B[k][j]
    # Adiciona a soma na linha
    linha.append(soma)
  # Adiciona a linha na matriz C
  C.append(linha)

# Imprime a matriz C, com os elementos separados por um espaÃ§o em branco
for i in range(m):
  print(*C[i])

'''
Exercicio 97 - Escreva um programa em Python que leia uma matriz de caracteres 
de dimensÃµes n x m e verifique se ela contÃ©m uma palavra escondida na horizontal, 
na vertical ou na diagonal. 
A entrada deve ser o valor de n, o valor de m, os elementos da matriz, 
um por linha, e a palavra a ser procurada. 
A saÃ­da deve ser â€œSIMâ€ se a palavra for encontrada, ou â€œNÃƒOâ€ caso contrÃ¡rio. Por exemplo:
Entrada:
4
5
A
B
C
D
E
F
G
H
I
J
K
L
M
N
O
P
Q
R
S
T
HI

Saida
SIM
'''

# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os caracteres
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = input()
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# LÃª a palavra a ser procurada
palavra = input()

# Cria uma variÃ¡vel booleana para indicar se a palavra foi encontrada ou nÃ£o
encontrada = False

# Verifica se a palavra estÃ¡ na horizontal
for i in range(n):
  # Concatena os elementos da i-Ã©sima linha em uma string
  linha = "".join(matriz[i])
  # Se a palavra for uma substring da linha, atualiza a variÃ¡vel encontrada
  if palavra in linha:
    encontrada = True

# Verifica se a palavra estÃ¡ na vertical
for j in range(m):
  # Cria uma lista vazia para armazenar os elementos da j-Ã©sima coluna
  coluna = []
  for i in range(n):
    # Adiciona o elemento da j-Ã©sima coluna e i-Ã©sima linha na lista
    coluna.append(matriz[i][j])
  # Concatena os elementos da lista em uma string
  coluna = "".join(coluna)
  # Se a palavra for uma substring da coluna, atualiza a variÃ¡vel encontrada
  if palavra in coluna:
    encontrada = True

# Verifica se a palavra estÃ¡ na diagonal principal
# Cria uma lista vazia para armazenar os elementos da diagonal principal
diagonal_principal = []
for i in range(n):
  for j in range(m):
    # Se i == j, o elemento pertence Ã  diagonal principal
    if i == j:
      # Adiciona o elemento na lista
      diagonal_principal.append(matriz[i][j])
# Concatena os elementos da lista em uma string
diagonal_principal = "".join(diagonal_principal)
# Se a palavra for uma substring da diagonal principal, atualiza a variÃ¡vel encontrada
if palavra in diagonal_principal:
  encontrada = True

# Verifica se a palavra estÃ¡ na diagonal secundÃ¡ria
# Cria uma lista vazia para armazenar os elementos da diagonal secundÃ¡ria
diagonal_secundaria = []
for i in range(n):
  for j in range(m):
    # Se i + j == n - 1, o elemento pertence Ã  diagonal secundÃ¡ria
    if i + j == n - 1:
      # Adiciona o elemento na lista
      diagonal_secundaria.append(matriz[i][j])
# Concatena os elementos da lista em uma string
diagonal_secundaria = "".join(diagonal_secundaria)
# Se a palavra for uma substring da diagonal secundÃ¡ria, atualiza a variÃ¡vel encontrada
if palavra in diagonal_secundaria:
  encontrada = True

# Imprime "SIM" se a palavra foi encontrada, ou "NÃƒO" caso contrÃ¡rio
if encontrada:
  print("SIM")
else:
  print("NÃƒO")



'''
Exercicio 98 - Escreva um programa em Python que leia uma matriz de nÃºmeros reais de 
dimensÃµes n x m e imprima a sua transposta. 
A transposta de uma matriz Ã© obtida trocando as linhas pelas colunas. 
A entrada deve ser o valor de n, o valor de m, e os elementos da matriz, um por linha. 
A saÃ­da deve ser a matriz transposta, com os elementos separados por um espaÃ§o em branco. 
Por exemplo:

Entrada:

2
3
1
2
3
4
5
6


SaÃ­da:

6 1 5
5 7 9

'''

# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria duas listas vazias para armazenar as somas das linhas e das colunas
soma_linhas = []
soma_colunas = []

# Calcula as somas das linhas
for i in range(n):
  # Inicializa a soma com zero
  soma = 0
  # Percorre os elementos da i-Ã©sima linha e acumula na soma
  for j in range(m):
    soma += matriz[i][j]
  # Adiciona a soma na lista soma_linhas
  soma_linhas.append(soma)

# Calcula as somas das colunas
for j in range(m):
  # Inicializa a soma com zero
  soma = 0
  # Percorre os elementos da j-Ã©sima coluna e acumula na soma
  for i in range(n):
    soma += matriz[i][j]
  # Adiciona a soma na lista soma_colunas
  soma_colunas.append(soma)

# Imprime as listas soma_linhas e soma_colunas, separadas por um espaÃ§o em branco
print(*soma_linhas)
print(*soma_colunas)


'''
Exercicio 99 - Escreva um programa em Python que leia uma matriz de nÃºmeros reais
de dimensÃµes n x m e imprima o maior e o menor elemento da matriz, 
bem como as suas posiÃ§Ãµes (linha e coluna). 
A entrada deve ser o valor de n, o valor de m, e os elementos da matriz, um por linha. 
A saÃ­da deve ser quatro nÃºmeros, separados por um espaÃ§o em branco, representando o
maior elemento, a sua linha, o seu coluna, e o menor elemento, a sua linha, a sua coluna. 
Por exemplo:

    Entrada
3
4
1.2
3.4
5.6
7.8
9.0
8.7
6.5
4.3
2.1
0.9
-1.0
-2.3

Saida
9.0 1 0 -2.3 2 3

'''

# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = float(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria quatro variÃ¡veis para armazenar o maior e o menor elemento da matriz, bem como as suas posiÃ§Ãµes
maior = matriz[0][0] # Inicializa o maior com o primeiro elemento da matriz
linha_maior = 0 # Inicializa a linha do maior com zero
coluna_maior = 0 # Inicializa a coluna do maior com zero
menor = matriz[0][0] # Inicializa o menor com o primeiro elemento da matriz
linha_menor = 0 # Inicializa a linha do menor com zero
coluna_menor = 0 # Inicializa a coluna do menor com zero

# Percorre a matriz e atualiza as variÃ¡veis
for i in range(n):
  for j in range(m):
    # Se o elemento da i-Ã©sima linha e j-Ã©sima coluna for maior que o maior, atualiza o maior e as suas posiÃ§Ãµes
    if matriz[i][j] > maior:
      maior = matriz[i][j]
      linha_maior = i
      coluna_maior = j
    # Se o elemento da i-Ã©sima linha e j-Ã©sima coluna for menor que o menor, atualiza o menor e as suas posiÃ§Ãµes
    if matriz[i][j] < menor:
      menor = matriz[i][j]
      linha_menor = i
      coluna_menor = j

# Imprime as variÃ¡veis, separadas por um espaÃ§o em branco
print(maior, linha_maior, coluna_maior, menor, linha_menor, coluna_menor)


'''
Exercicio 100 - Escreva um programa em Python que leia uma matriz de nÃºmeros inteiros 
de dimensÃµes n x m e imprima a matriz rotacionada 90 graus no sentido horÃ¡rio. 
A entrada deve ser o valor de n, o valor de m, e os elementos da matriz, um por linha. 
A saÃ­da deve ser a matriz rotacionada, com os elementos separados por um espaÃ§o em branco. 
Por exemplo:
Entrada:
3
3
1
2
3
4
5
6
7
8
9

Saida
7 4 1
8 5 2
9 6 3

'''

# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria uma matriz vazia para armazenar os elementos da matriz rotacionada
rotacionada = []

# Calcula a matriz rotacionada
for j in range(m):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for i in range(n - 1, -1, -1):
    # Adiciona o elemento da i-Ã©sima linha e j-Ã©sima coluna na linha
    linha.append(matriz[i][j])
  # Adiciona a linha na matriz rotacionada
  rotacionada.append(linha)

# Imprime a matriz rotacionada, com os elementos separados por um espaÃ§o em branco
for i in range(m):
  print(*rotacionada[i])


'''
Exercicio 101 - Escreva um programa em Python que leia uma matriz de nÃºmeros 
inteiros de dimensÃµes n x m e imprima a matriz espelhada horizontalmente. 
A matriz espelhada horizontalmente Ã© obtida invertendo a ordem das colunas da 
matriz original. A entrada deve ser o valor de n, o valor de m, 
e os elementos da matriz, um por linha. A saÃ­da deve ser a matriz espelhada, 
com os elementos separados por um espaÃ§o em branco. Por exemplo:
    
Entradas:
2
3
1
2
3
4
5
6
    
Saida
3 2 1
6 5 4

'''

# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria uma matriz vazia para armazenar os elementos da matriz espelhada
espelhada = []

# Calcula a matriz espelhada
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m - 1, -1, -1):
    # Adiciona o elemento da i-Ã©sima linha e j-Ã©sima coluna na linha
    linha.append(matriz[i][j])
  # Adiciona a linha na matriz espelhada
  espelhada.append(linha)

# Imprime a matriz espelhada, com os elementos separados por um espaÃ§o em branco
for i in range(n):
  print(*espelhada[i])
  
        
'''
Exercicio 102 - Escreva um programa em Python que leia uma matriz de nÃºmeros inteiros
de dimensÃµes n x m e imprima a matriz espelhada verticalmente. 
A matriz espelhada verticalmente Ã© obtida invertendo a ordem das linhas da matriz
original. A entrada deve ser o valor de n, o valor de m, e os elementos da matriz, 
um por linha. A saÃ­da deve ser a matriz espelhada, com os elementos separados 
por um espaÃ§o em branco. Por exemplo:
Entrdas
2
3
1
2
3
4
5
6

Saida
4 5 6
1 2 3
    
'''
# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria uma matriz vazia para armazenar os elementos da matriz espelhada
espelhada = []

# Calcula a matriz espelhada
for i in range(n - 1, -1, -1):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # Adiciona o elemento da i-Ã©sima linha e j-Ã©sima coluna na linha
    linha.append(matriz[i][j])
  # Adiciona a linha na matriz espelhada
  espelhada.append(linha)

# Imprime a matriz espelhada, com os elementos separados por um espaÃ§o em branco
for i in range(n):
  print(*espelhada[i])

'''
Exercicio 103 - Escreva um programa em Python que leia uma matriz de nÃºmeros inteiros
de dimensÃµes n x m e imprima a matriz ordenada. 
A matriz ordenada Ã© obtida colocando os elementos da matriz original em ordem crescente, 
da esquerda para a direita e de cima para baixo. 
A entrada deve ser o valor de n, o valor de m, e os elementos da matriz, um por linha. 
A saÃ­da deve ser a matriz ordenada, com os elementos separados por um espaÃ§o em branco. 
Por exemplo:
Entrada
3
3
9
8
7
6
5
4
3
2
1

Saida
1 2 3
4 5 6
7 8 9
   
'''
# LÃª os valores de n e m
n = int(input())
m = int(input())

# Cria uma matriz vazia para armazenar os elementos da matriz
matriz = []

# LÃª os elementos da matriz
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # LÃª um elemento e adiciona na linha
    elemento = int(input())
    linha.append(elemento)
  # Adiciona a linha na matriz
  matriz.append(linha)

# Cria uma lista vazia para armazenar os elementos da matriz em uma Ãºnica dimensÃ£o
lista = []

# Percorre a matriz e adiciona os elementos na lista
for i in range(n):
  for j in range(m):
    lista.append(matriz[i][j])

# Ordena a lista em ordem crescente
lista.sort()

# Cria uma matriz vazia para armazenar os elementos da matriz ordenada
ordenada = []

# Percorre a lista e distribui os elementos na matriz ordenada
k = 0 # Ãndice para percorrer a lista
for i in range(n):
  # Cria uma lista vazia para representar uma linha
  linha = []
  for j in range(m):
    # Adiciona o elemento da lista na linha
    linha.append(lista[k])
    # Incrementa o Ã­ndice da lista
    k += 1
  # Adiciona a linha na matriz ordenada
  ordenada.append(linha)

# Imprime a matriz ordenada, com os elementos separados por um espaÃ§o em branco
for i in range(n):
  print(*ordenada[i])

'''
Exercicios de manipulacao de aquivos
'''


'''
Exercicio 104 - ler conteudo do arquivo

'''
 # ler arquivo de texto
arquivo = open('nomes.txt', 'r', encoding='utf-8')    # r = read

# lÃª todo o conteÃºdo do arquivo
texto = arquivo.read()

lista = texto.split('\n')
print(texto)

print(lista)

# fecha o arquivo
arquivo.close()


'''
Exercicio 105 - ler conteudo do arquivo

'''

# ler arquivo de texto linha por linha

arquivo = open('numeros.txt', 'r', encoding='utf-8')

lista = []

for linha in arquivo:
    print(linha, end='')
    lista.append(int(linha))

print(lista)

arquivo.close()


'''
Exercicio 106 - criar arquivo de texto

'''

arquivo = open('meu arquivo.txt', 'w')

arquivo.write('exemplo de texto\n')
arquivo.write('nova linha\n')
arquivo.write('alguma outra coisa\n')

nome = input('Digite um nome: ')
arquivo.write(nome + '\n')

idade = int(input('Digite a idade: '))
arquivo.write(str(idade) + '\n')

arquivo.close()

'''
Exercicio 107 - Cadastra

'''

def cadastrar():
    arquivo = open('nomes.txt', 'a', encoding='utf-8')        # a = append
    while True:
        nome = input('Digite um nome (0 para finalizar): ')
        if nome == '0':
            break
        arquivo.write(nome + '\n')
    arquivo.close()

def exibir():
    arquivo = open('nomes.txt', 'r', encoding='utf-8')
    for linha in arquivo:
        print(linha, end='')
    arquivo.close()

def excluir():
    nome = input('Digite o nome para excluir: ')

    arquivo = open('nomes.txt', 'r', encoding='utf-8')
    lista = []
    for linha in arquivo:
        lista.append(linha.replace('\n', ''))
    arquivo.close()

    if nome in lista:
        lista.remove(nome)

    arquivo = open('nomes.txt', 'w', encoding='utf-8')
    for nome in lista:
        arquivo.write(nome + '\n')
    arquivo.close()

while True:
    print()
    print('1 - Cadastrar')
    print('2 - Exibir')
    print('3 - Excluir')
    print('4 - Sair')
    opcao = int(input('Selecione a opÃ§Ã£o: '))
    match opcao:
        case 1:
            cadastrar()
        case 2:
            exibir()
        case 3:
            excluir()
        case 4:
            break
        case _:
            print('OpÃ§Ã£o invÃ¡lida')


'''
Exercicio 108 - gerar um arquivo com 100,000 de linhas e ler

'''

# gerar um arquivo com 1 milha de linhas e ler

arquivo = open('novo arquivo.txt', 'w', encoding='utf-8')
for i in range(100000):
    arquivo.write('exemplo de texto\n')
arquivo.close()

arquivo = open('novo arquivo.txt', 'r', encoding='utf-8')
quantidade_linhas = 0
for linha in arquivo:
    # print(f'{quantidade_linhas} {linha}')
    quantidade_linhas += 1

print(f'Quantidade de linhas: {quantidade_linhas}')
arquivo.close()

'''
ManipulaÃ§Ã£o de arquivos CSV

'''

'''
Exercicio 109 - leitura de um aquivo CSV

'''

import csv

with open('dados.csv', 'r') as arquivo:
    leitor_csv = csv.reader(arquivo)
    for linha in leitor_csv:
        print(linha)


'''
Exercicio 110 - Filtrando linhas em um arquivo CSV

'''

import csv

with open('dados.csv', 'r') as arquivo:
    leitor_csv = csv.reader(arquivo)
    next(leitor_csv)  # pula a primeira linha (cabecalho)
    for linha in leitor_csv:
        if linha[2] == 'Brasil':
            print(linha)


'''
Exercicio 111 - SeleÃ§Ã£o de colunas especÃ­ficas no arquivo csv

'''

import csv

with open('dados.csv', 'r') as arquivo:
    leitor_csv = csv.reader(arquivo)
    for linha in leitor_csv:
        print(linha[0], linha[2])
        

'''
Exercicio 112 - Escrevendo um arquivo CSV

'''


import csv

dados = [
    ['Nome', 'Idade', 'PaÃ­s'],
    ['Joao', '25', 'Brasil'],
    ['Maria', '30', 'Portugal'],
    ['João', '35', 'Espanha']
]

with open('dados.csv', 'w', newline='') as arquivo:
    escritor_csv = csv.writer(arquivo)
    for linha in dados:
        escritor_csv.writerow(linha)

'''
Exercicio 113 - Crie um arquivo CSV vazio chamado "dados.csv".

'''

import csv

with open('dados.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerow([])


'''
Exercicio 114 - Crie um arquivo CSV chamado "alunos.csv" e escreva 
dados de alunos nele.

'''

import csv

dados_alunos = [
    ['Nome', 'Idade', 'Nota'],
    ['Alice', 20, 95],
    ['Bob', 22, 88],
    ['Carol', 21, 92]
]

with open('alunos.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerows(dados_alunos)

'''
Exercicio 115 -  Leia dados do arquivo "alunos.csv" e imprima-os na tela.

'''

import csv

with open('alunos.csv', 'r') as arquivo_csv:
    reader = csv.reader(arquivo_csv)
    for linha in reader:
        print(linha)


'''
Exercicio 116 - Adicione um novo aluno ao arquivo "alunos.csv".

'''

import csv

novo_aluno = ['David', 23, 89]

with open('alunos.csv', 'a', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerow(novo_aluno)


'''
Exercicio 117 - Atualize a nota de "Bob" para 90 no arquivo "alunos.csv".

'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv:
    linhas = list(csv.reader(arquivo_csv))

for linha in linhas:
    if linha[0] == 'Bob':
        linha[2] = 90

with open('alunos.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerows(linhas)

'''
Exercicio 118 - Remova a linha correspondente a "Carol" do 
arquivo "alunos.csv".

'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv:
    linhas = list(csv.reader(arquivo_csv))

for linha in linhas:
    if linha[0] == 'Carol':
        linhas.remove(linha)

with open('alunos.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerows(linhas)

'''
Exercicio 119 - Calcule a mÃ©dia das notas no arquivo "alunos.csv".

'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv:
    reader = csv.reader(arquivo_csv)
    next(reader)  # Pula o cabecalho

    notas = [float(linha[2]) for linha in reader]

media_notas = sum(notas) / len(notas)
print(f'MÃ©dia das notas: {media_notas:.2f}')

'''
Exercicio 120 - Crie um novo arquivo "alunos_aprovados.csv" 
com apenas os alunos que tÃªm notas acima de 90.

'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv:
    reader = csv.reader(arquivo_csv)
    cabecalho = next(reader)

    alunos_aprovados = [linha for linha in reader if float(linha[2]) > 90]

with open('alunos_aprovados.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerow(cabecalho)
    writer.writerows(alunos_aprovados)


'''
Exercicio 121 - Concatenar dois arquivos CSV:
    
'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv1, \
     open('alunos_aprovados.csv', 'r', newline='') as arquivo_csv2:
    reader1 = csv.reader(arquivo_csv1)
    reader2 = csv.reader(arquivo_csv2)
    cabecalho = next(reader1)
    alunos_aprovados = list(reader2)
    todos_alunos = list(reader1)

with open('todos_alunos.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerow(cabecalho)
    writer.writerows(alunos_aprovados + todos_alunos)

'''
Exercicio 122 - Ordenar um arquivo CSV por uma coluna:
Ordene o arquivo "alunos.csv" pelo nome dos alunos em ordem alfabÃ©tica e salve o resultado em 
"alunos_ordenados.csv".

'''

import csv

with open('alunos.csv', 'r', newline='') as arquivo_csv:
    reader = csv.reader(arquivo_csv)
    cabecalho = next(reader)
    linhas = list(reader)

linhas_ordenadas = sorted(linhas, key=lambda x: x[0])

with open('alunos_ordenados.csv', 'w', newline='') as arquivo_csv:
    writer = csv.writer(arquivo_csv)
    writer.writerow(cabecalho)
    writer.writerows(linhas_ordenadas)


'''
ManipulaÃ§Ã£o de arquivos Excel (e automaÃ§Ã£o)

'''

'''
Exercicio 123 - Criando uma planilha nova 

'''

# inicialmente NO TERMINAL  use: pip install openpyxl

from openpyxl import Workbook

# Cria uma nova planilha em branco
wb = Workbook()

# Seleciona a planilha ativa
planilha = wb.active

# Define o valor da cÃ©lula A1
planilha['A1'] = 'OlÃ¡, Excel com Python!'

# Salva o arquivo
wb.save('Exemplo.xlsx') 

'''
Exercicio 124 - Lendo dados de uma planilha existente

'''

from openpyxl import load_workbook

# Carrega o arquivo existente
wb = load_workbook('Exemplo.xlsx')

# Seleciona a planilha ativa
planilha = wb.active

# LÃª o valor da cÃ©lula A1
valor_celula = planilha['A1'].value

# Imprime o valor na tela
print(valor_celula)


'''
Exercicio 125 - Atualizando dados em uma planilha existente

'''

from openpyxl import load_workbook

# Carrega o arquivo existente
wb = load_workbook('Exemplo.xlsx')

# Seleciona a planilha ativa
planilha = wb.active

# Atualiza o valor da cÃ©lula A1
planilha['A1'] = 'Novo valor'

# Salva o arquivo
wb.save('Exemplo.xlsx')


'''
Exercicio 126 - Crie um programa que leia um arquivo CSV chamado 
"dados.csv" e calcule a mÃ©dia de uma coluna especÃ­fica, como a coluna "nota".


'''

import csv

def calcular_media(coluna):
    total = 0
    contador = 0

    with open("dados.csv", "r") as arquivo_csv:
        leitor_csv = csv.DictReader(arquivo_csv)
        for linha in leitor_csv:
            valor = float(linha[coluna])
            total += valor
            contador += 1

    if contador > 0:
        media = total / contador
        return media
    else:
        return 0

coluna_desejada = "nota"
media = calcular_media(coluna_desejada)
print(f"A mÃ©dia da coluna '{coluna_desejada}' Ã© {media:.2f}")


'''
Exercicio 127 - Crie um programa que leia um arquivo CSV chamado "dados.csv", 
ordene as linhas com base em uma coluna especÃ­fica, como a coluna "nome", 
e salve o resultado em um novo arquivo chamado "dados_ordenados.csv".

'''

import csv

def ordenar_por_coluna(input_file, output_file, coluna):
    with open(input_file, "r") as arquivo_csv:
        leitor_csv = csv.DictReader(arquivo_csv)
        linhas = sorted(leitor_csv, key=lambda linha: linha[coluna])

    with open(output_file, "w", newline="") as arquivo_saida:
        escritor_csv = csv.DictWriter(arquivo_saida, fieldnames=linhas[0].keys())
        escritor_csv.writeheader()
        escritor_csv.writerows(linhas)

ordenar_por_coluna("dados.csv", "dados_ordenados.csv", "nome")


'''
Exercicio 128 - Crie um programa que leia dois arquivos CSV chamados "dados1.csv" 
e "dados2.csv" e mescle-os em um Ãºnico arquivo chamado "dados_mesclados.csv", 
eliminando duplicatas com base em uma coluna comum, como a coluna "id".
 '''
 
import csv

def mesclar_arquivos(input_file1, input_file2, output_file, coluna_comum):
    dados_mesclados = []

    with open(input_file1, "r") as arquivo1, open(input_file2, "r") as arquivo2:
        leitor_csv1 = csv.DictReader(arquivo1)
        leitor_csv2 = csv.DictReader(arquivo2)
        dados1 = list(leitor_csv1)
        dados2 = list(leitor_csv2)

    for linha1 in dados1:
        for linha2 in dados2:
            if linha1[coluna_comum] == linha2[coluna_comum]:
                dados_mesclados.append({**linha1, **linha2})
                break

    with open(output_file, "w", newline="") as arquivo_saida:
        fieldnames = dados_mesclados[0].keys()
        escritor_csv = csv.DictWriter(arquivo_saida, fieldnames=fieldnames)
        escritor_csv.writeheader()
        escritor_csv.writerows(dados_mesclados)

mesclar_arquivos("dados1.csv", "dados2.csv", "dados_mesclados.csv", "id")

 
'''
Exercicio 129 - Crie um programa que leia uma planilha Excel chamada "dados.xlsx" e 
crie um grÃ¡fico com base em seus dados, como um grÃ¡fico de barras 
representando uma coluna especÃ­fica.

'''

import openpyxl
from openpyxl.chart import BarChart, Reference

def criar_grafico(input_file, coluna_dados, output_file):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    chart = BarChart()
    chart.title = "GrÃ¡fico de Barras"
    chart.x_axis.title = "Categorias"
    chart.y_axis.title = "Valores"

    data = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    sheet.add_chart(chart, "E5")

    workbook.save(output_file)

criar_grafico("dados.xlsx", "B", "grafico.xlsx")


'''
Exercicio 130 - Crie um programa que leia vÃ¡rios arquivos CSV com informaÃ§Ãµes de 
vendas de diferentes produtos e gere um relatÃ³rio consolidado em um 
novo arquivo chamado "relatorio_vendas.csv", incluindo informaÃ§Ãµes como o total 
de vendas por produto.

'''

import csv
import os

def gerar_relatorio(input_folder, output_file):
    relatorio = {}
    for arquivo in os.listdir(input_folder):
        if arquivo.endswith(".csv"):
            with open(os.path.join(input_folder, arquivo), "r") as arquivo_csv:
                leitor_csv = csv.DictReader(arquivo_csv)
                for linha in leitor_csv:
                    produto = linha["produto"]
                    valor_venda = float(linha["valor"])
                    if produto in relatorio:
                        relatorio[produto] += valor_venda
                    else:
                        relatorio[produto] = valor_venda

    with open(output_file, "w", newline="") as arquivo_saida:
        escritor_csv = csv.writer(arquivo_saida)
        escritor_csv.writerow(["Produto", "Total de Vendas"])
        for produto, total_vendas in relatorio.items():
            escritor_csv.writerow([produto, total_vendas])

gerar_relatorio("dados_vendas", "relatorio_vendas.csv")


'''
Exercicio 131 - Crie um programa que leia um arquivo de texto chamado "secreto.txt", 
criptografe seu conteÃºdo usando um algoritmo de criptografia de sua escolha e
 salve o resultado em um novo arquivo chamado "secreto_criptografado.txt".
 
 '''
 

# Implemente um algoritmo de criptografia de sua escolha
# Exemplo simples de substituiÃ§Ã£o de caracteres:

def criptografar(texto):
    chave = {"a": "x", "b": "y", "c": "z", "d": "a", "e": "b", "f": "c", "g": "d", "h": "e", "i": "f", "j": "g", "k": "h", "l": "i", "m": "j", "n": "k", "o": "l", "p": "m", "q": "n", "r": "o", "s": "p", "t": "q", "u": "r", "v": "s", "w": "t", "x": "u", "y": "v", "z": "w", " ": " "}
    texto_criptografado = "".join([chave.get(letra, letra) for letra in texto])
    return texto_criptografado

with open("secreto.txt", "r") as arquivo:
    conteudo = arquivo.read()
    conteudo_criptografado = criptografar(conteudo)

with open("secreto_criptografado.txt", "w") as arquivo_cripto:
    arquivo_cripto.write(conteudo_criptografado)
 
    
'''
Exercicio 132 - Crie um programa que leia um arquivo criptografado chamado
 "secreto_criptografado.txt", decodifique seu conteÃºdo e salve o resultado
 em um novo arquivo chamado "secreto_decodificado.txt".

'''

# Implemente a função de decodificaÃ§Ã£o correspondente ao algoritmo de criptografia

def decodificar(texto_criptografado):
    chave = {"x": "a", "y": "b", "z": "c", "a": "d", "b": "e", "c": "f", "d": "g", "e": "h", "f": "i", "g": "j", "h": "k", "i": "l", "j": "m", "k": "n", "l": "o", "m": "p", "n": "q", "o": "r", "p": "s", "q": "t", "r": "u", "s": "v", "t": "w", "u": "x", "v": "y", "w": "z", " ": " "}
    texto_decodificado = "".join([chave.get(letra, letra) for letra in texto_criptografado])
    return texto_decodificado

with open("secreto_criptografado.txt", "r") as arquivo_cripto:
    conteudo_criptografado = arquivo_cripto.read()
    conteudo_decodificado = decodificar(conteudo_criptografado)

with open("secreto_decodificado.txt", "w") as arquivo_decodificado:
    arquivo_decodificado.write(conteudo_decodificado)

'''
Exercicio 133 -  Crie um programa que leia um arquivo de texto chamado "longo.txt", 
aplique um algoritmo de compressÃ£o de sua escolha e salve o resultado em um novo 
arquivo chamado "longo_comprimido.txt".

'''

import zlib

with open("longo.txt", "r") as arquivo:
    conteudo = arquivo.read()
    conteudo_comprimido = zlib.compress(conteudo.encode("utf-8"))

with open("longo_comprimido.txt", "wb") as arquivo_comprimido:
    arquivo_comprimido.write(conteudo_comprimido)

'''
Exercicio 134 - Crie um programa que leia um arquivo comprimido chamado 
"longo_comprimido.txt", aplique a descompressÃ£o correspondente e salve o resultado 
em um novo arquivo chamado "longo_descomprimido.txt".

'''

import zlib

with open("longo_comprimido.txt", "rb") as arquivo_comprimido:
    conteudo_comprimido = arquivo_comprimido.read()
    conteudo_descomprimido = zlib.decompress(conteudo_comprimido).decode("utf-8")

with open("longo_descomprimido.txt", "w") as arquivo_descomprimido:
    arquivo_descomprimido.write(conteudo_descomprimido)

'''
Exercicio 134 - Crie um programa que leia uma planilha Excel chamada
"planilha_complexa.xlsx" com mÃºltiplas abas, realize operaÃ§Ãµes avanÃ§adas, 
como a criaÃ§Ã£o de grÃ¡ficos, formataÃ§Ã£o condicional e adiÃ§Ã£o de fÃ³rmulas, e 
salve o resultado em um novo arquivo chamado "planilha_resultado.xlsx".

'''

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference

def manipular_planilha(input_file, output_file):
    workbook = openpyxl.load_workbook(input_file)

    # Adicione cÃ³digo aqui para manipular a planilha, criar grÃ¡ficos e aplicar formataÃ§Ã£o condicional

    workbook.save(output_file)

manipular_planilha("planilha_complexa.xlsx", "planilha_resultado.xlsx")

'''
Exercicio 135 -  VocÃª recebeu uma planilha Excel chamada "vendas.xlsx" que 
contÃ©m dados de vendas de produtos em diferentes regiÃµes. Crie um programa em 
Python que leia essa planilha, calcule o total de vendas para cada produto e 
regiÃ£o, e gere um novo arquivo Excel chamado "relatorio_vendas.xlsx" com os resultados.

'''

import openpyxl

# Abra a planilha de vendas
workbook = openpyxl.load_workbook("vendas.xlsx")
sheet = workbook.active

# Crie um dicionÃ¡rio para armazenar os totais de vendas por produto e regiÃ£o
relatorio = {}

# Itere pelas linhas da planilha
for row in sheet.iter_rows(min_row=2, values_only=True):
    produto, regiao, valor = row
    chave = (produto, regiao)
    
    if chave in relatorio:
        relatorio[chave] += valor
    else:
        relatorio[chave] = valor

# Crie um novo arquivo Excel para o relatÃ³rio
novo_workbook = openpyxl.Workbook()
novo_sheet = novo_workbook.active

# Preencha o novo arquivo Excel com os dados do relatÃ³rio
for chave, total in relatorio.items():
    produto, regiao = chave
    novo_sheet.append([produto, regiao, total])

# Salve o arquivo do relatÃ³rio
novo_workbook.save("relatorio_vendas.xlsx")

'''
Exercicio 136 - VocÃª possui uma planilha Excel chamada "estoque.xlsx" que contÃ©m 
informaÃ§Ãµes sobre o estoque de produtos em uma loja. Crie um programa em Python 
que leia uma lista de novos produtos e suas quantidades em um arquivo CSV chamado 
novos_produtos.csv" e atualize a planilha "estoque.xlsx" com esses novos dados.    

'''

import openpyxl
import csv

# Abra a planilha de estoque
workbook = openpyxl.load_workbook("estoque.xlsx")
sheet = workbook.active

# Leia os novos produtos do arquivo CSV
novos_produtos = []
with open("novos_produtos.csv", "r") as arquivo_csv:
    leitor_csv = csv.reader(arquivo_csv)
    for linha in leitor_csv:
        produto, quantidade = linha
        novos_produtos.append((produto, quantidade))

# Atualize a planilha com os novos dados
for produto, quantidade in novos_produtos:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        produto_atual, estoque_atual = row
        if produto == produto_atual:
            sheet.cell(row=row[0], column=2, value=int(estoque_atual) + int(quantidade))
            break
    else:
        # Se o produto nÃ£o existir na planilha, adicione-o
        sheet.append([produto, quantidade])

# Salve a planilha atualizada
workbook.save("estoque_atualizado.xlsx")

'''
______________________________________________________________________________________

FIM DO CURSO DE PYTHON - 2023 - ENGENHARIA DE SOFTWARE - PROF EDER
______________________________________________________________________________________

'''

